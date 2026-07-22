"""Master Pegawai — data kepegawaian menyeluruh satker (adopsi SIMAN-G).

BERBEDA dari Referensi Pejabat (khusus pejabat penatausahaan/penanda tangan):
master ini menampung SELURUH pegawai + unit kerjanya, sebagai rujukan lintas
modul. Semua user login dapat melihat; admin mengelola (CRUD). NIP unik bila
diisi. Pola sama dengan referensi pejabat/ruangan.
"""
import uuid
from datetime import datetime, timezone
from typing import Optional

from fastapi import (APIRouter, Depends, File, Form, HTTPException, Request,
                     Response, UploadFile)
from pydantic import BaseModel, Field

from auth_utils import require_admin, require_user, require_user_or_query_token
from db import db, fs_bucket
from kartu_utils import hash_kandidat, label_kartu, valid_uid
from shared_utils import (kode_satker_user, limiter, log_audit,
                          pastikan_akses_dok_satker, scope_query_field_satker)
from pejabat_utils import JENIS_PELAKSANA, STATUS_KEPEGAWAIAN
from pegawai_utils import (
    AGAMA, DIGIT_BANK, JENIS_IDENTITAS_WNA, JENIS_JABATAN, JENIS_KELAMIN,
    JENIS_KONTRAK_NON_ASN, KATEGORI_PEGAWAI, KEWARGANEGARAAN,
    PANGKAT_GOLONGAN, STATUS_PEGAWAI,
    STATUS_PERKAWINAN, SUB_KATEGORI_NON_ASN, baris_impor_ke_pegawai,
    beda_snapshot_pemegang, deteksi_identitas, info_masa_pegawai,
    kelompok_unit_kerja, pegawai_perlu_serah_terima, rekap_eselon,
    snapshot_pemegang_aset, validate_pegawai,
)

pegawai_router = APIRouter()

# kartu_uid_hashes TIDAK pernah dikirim ke klien (cukup kartu_label utk UI) —
# hash HMAC memang tak bisa dibalik, tapi tak ada alasan membocorkannya.
# kartu_terdaftar_oleh (username admin) juga internal — jejaknya di audit log.
_PROJ = {"_id": 0, "kartu_uid_hashes": 0, "kartu_terdaftar_oleh": 0}


class PegawaiIn(BaseModel):
    nama: str
    nip: Optional[str] = ""
    gelar_depan: Optional[str] = ""
    gelar_belakang: Optional[str] = ""
    kewarganegaraan: Optional[str] = ""
    jenis_identitas_wna: Optional[str] = ""
    nomor_identitas_wna: Optional[str] = ""
    jenis_kelamin: Optional[str] = ""
    tempat_lahir: Optional[str] = ""
    tanggal_lahir: Optional[str] = ""
    agama: Optional[str] = ""
    status_perkawinan: Optional[str] = ""
    status_kepegawaian: Optional[str] = ""
    sub_kategori_non_asn: Optional[str] = ""
    pangkat_golongan: Optional[str] = ""
    jabatan: Optional[str] = ""
    # Rangkap jabatan struktural sementara (Plt/Plh): jenis_pelaksana
    # ""/plt/plh + jabatan yang di-Plt/Plh-kan (jabatan definitif tetap di
    # field `jabatan`).
    jenis_pelaksana: Optional[str] = ""
    jabatan_pelaksana: Optional[str] = ""
    jenis_jabatan: Optional[str] = ""
    kategori_pegawai: Optional[str] = ""
    eselon: Optional[str] = ""
    eselon1: Optional[str] = ""
    eselon2: Optional[str] = ""
    eselon3: Optional[str] = ""
    eselon4: Optional[str] = ""
    eselon5: Optional[str] = ""
    unit_kerja: Optional[str] = ""
    unit_organisasi: Optional[str] = ""
    npwp: Optional[str] = ""
    pendidikan_terakhir: Optional[str] = ""
    no_hp: Optional[str] = ""
    email: Optional[str] = ""
    alamat: Optional[str] = ""
    nama_bank: Optional[str] = ""
    no_rekening: Optional[str] = ""
    nomor_kontrak: Optional[str] = ""
    tgl_mulai_kontrak: Optional[str] = ""
    tgl_selesai_kontrak: Optional[str] = ""
    # Non-ASN: kontrak internal instansi vs OUTSOURCING + perusahaan penyedia
    jenis_kontrak_non_asn: Optional[str] = ""
    perusahaan_penyedia: Optional[str] = ""
    tmt_jabatan: Optional[str] = ""
    # Akhir periode jabatan (utk hitung sisa masa jabatan di daftar)
    tanggal_akhir_jabatan: Optional[str] = ""
    # Penghubung lintas modul: kode satker (6 digit) + kode lengkap 12 digit
    kode_satker: Optional[str] = ""
    kode_satker_lengkap: Optional[str] = ""
    status: Optional[str] = "aktif"
    keterangan: Optional[str] = ""


def _bersih(p: PegawaiIn) -> dict:
    doc = {
        "nama": str(p.nama or "").strip(),
        "nip": str(p.nip or "").strip(),
        "gelar_depan": str(p.gelar_depan or "").strip(),
        "gelar_belakang": str(p.gelar_belakang or "").strip(),
        "kewarganegaraan": str(p.kewarganegaraan or "").strip().lower(),
        "jenis_identitas_wna": str(p.jenis_identitas_wna or "").strip().lower(),
        "nomor_identitas_wna": str(p.nomor_identitas_wna or "").strip(),
        "jenis_kelamin": str(p.jenis_kelamin or "").strip().upper(),
        "tempat_lahir": str(p.tempat_lahir or "").strip(),
        "tanggal_lahir": str(p.tanggal_lahir or "").strip()[:10],
        "agama": str(p.agama or "").strip(),
        "status_perkawinan": str(p.status_perkawinan or "").strip(),
        "status_kepegawaian": str(p.status_kepegawaian or "").strip(),
        "sub_kategori_non_asn": str(p.sub_kategori_non_asn or "").strip(),
        "pangkat_golongan": str(p.pangkat_golongan or "").strip(),
        "jabatan": str(p.jabatan or "").strip(),
        "jenis_pelaksana": str(p.jenis_pelaksana or "").strip().lower(),
        "jabatan_pelaksana": str(p.jabatan_pelaksana or "").strip(),
        "jenis_jabatan": str(p.jenis_jabatan or "").strip(),
        "kategori_pegawai": str(p.kategori_pegawai or "").strip(),
        "eselon": str(p.eselon or "").strip(),
        "eselon1": str(p.eselon1 or "").strip(),
        "eselon2": str(p.eselon2 or "").strip(),
        "eselon3": str(p.eselon3 or "").strip(),
        "eselon4": str(p.eselon4 or "").strip(),
        "eselon5": str(p.eselon5 or "").strip(),
        "unit_kerja": str(p.unit_kerja or "").strip(),
        "unit_organisasi": str(p.unit_organisasi or "").strip(),
        "npwp": str(p.npwp or "").strip(),
        "pendidikan_terakhir": str(p.pendidikan_terakhir or "").strip(),
        "no_hp": str(p.no_hp or "").strip(),
        "email": str(p.email or "").strip(),
        "alamat": str(p.alamat or "").strip(),
        "nama_bank": str(p.nama_bank or "").strip(),
        "no_rekening": str(p.no_rekening or "").strip(),
        "nomor_kontrak": str(p.nomor_kontrak or "").strip(),
        "tgl_mulai_kontrak": str(p.tgl_mulai_kontrak or "").strip()[:10],
        "tgl_selesai_kontrak": str(p.tgl_selesai_kontrak or "").strip()[:10],
        "jenis_kontrak_non_asn": str(p.jenis_kontrak_non_asn or "").strip().lower(),
        "perusahaan_penyedia": str(p.perusahaan_penyedia or "").strip(),
        "tmt_jabatan": str(p.tmt_jabatan or "").strip()[:10],
        "tanggal_akhir_jabatan": str(p.tanggal_akhir_jabatan or "").strip()[:10],
        "kode_satker": str(p.kode_satker or "").strip(),
        "kode_satker_lengkap": str(p.kode_satker_lengkap or "").strip(),
        "status": str(p.status or "aktif").strip() or "aktif",
        "keterangan": str(p.keterangan or "").strip(),
    }
    # Unit kerja efektif = Eselon terdalam bila field unit_kerja kosong (agar
    # rekap & tampilan tetap punya satu label unit meski data berjenjang).
    if not doc["unit_kerja"]:
        from pegawai_utils import unit_kerja_terdalam
        doc["unit_kerja"] = unit_kerja_terdalam(doc)
    return doc


async def _nip_bentrok(nip: str, kecuali_id: str = "", kode: str = "") -> bool:
    """True bila NIP sudah dipakai pegawai LAIN di satker sama (NIP kosong tak
    pernah bentrok). Dibatasi satker agar admin satker A tak terganggu data
    satker B yang tak terlihat olehnya."""
    if not nip:
        return False
    q = {"nip": nip}
    if kode:
        q["kode_satker"] = {"$in": [kode, "", None]}
    if kecuali_id:
        q["id"] = {"$ne": kecuali_id}
    return await db.pegawai.find_one(q, _PROJ) is not None


@pegawai_router.get("/pegawai/referensi")
async def referensi_pegawai(_user: dict = Depends(require_user)):
    """Referensi pilihan (untuk dropdown UI)."""
    def _opt(d):
        return [{"kode": k, "uraian": v} for k, v in d.items()]
    return {
        "jenis_kelamin": _opt(JENIS_KELAMIN),
        "status_kepegawaian": _opt(STATUS_KEPEGAWAIAN),
        "jenis_pelaksana": _opt(JENIS_PELAKSANA),
        "jenis_jabatan": _opt(JENIS_JABATAN),
        "kategori_pegawai": _opt(KATEGORI_PEGAWAI),
        "sub_kategori_non_asn": _opt(SUB_KATEGORI_NON_ASN),
        "agama": _opt(AGAMA),
        "status_perkawinan": _opt(STATUS_PERKAWINAN),
        "status": _opt(STATUS_PEGAWAI),
        "kewarganegaraan": _opt(KEWARGANEGARAAN),
        "jenis_identitas_wna": _opt(JENIS_IDENTITAS_WNA),
        "jenis_kontrak_non_asn": _opt(JENIS_KONTRAK_NON_ASN),
        # saran pangkat MENGIKUTI status kepegawaian + peta digit bank utk
        # peringatan lunak rekening (pola form KERJA-BARENG).
        "pangkat_golongan": PANGKAT_GOLONGAN,
        "digit_bank": DIGIT_BANK,
    }


@pegawai_router.get("/pegawai/rekap-unit")
async def rekap_unit_pegawai(_user: dict = Depends(require_user)):
    """Rekap jumlah pegawai per unit kerja & per Eselon I (untuk ringkasan)."""
    semua = await db.pegawai.find(
        scope_query_field_satker(_user), _PROJ).to_list(20000)
    kelompok = kelompok_unit_kerja(semua)
    return {
        "unit": [{"unit_kerja": g["unit_kerja"], "jumlah": g["jumlah"]} for g in kelompok],
        "eselon1": rekap_eselon(semua, "eselon1"),
        "jumlah_pegawai": len(semua),
        "jumlah_unit": len(kelompok),
    }


@pegawai_router.get("/pegawai")
async def list_pegawai(_user: dict = Depends(require_user)):
    """Daftar seluruh pegawai satker (terurut nama).

    Tiap item diperkaya `info_masa` (identitas terdeteksi NIP/NI PPPK/NRP/
    NIK, perkiraan pensiun per BUP UU 20/2023 / UU TNI 3/2025 / UU Polri
    5/2026, sisa masa jabatan & kontrak) — bahan kolom durasi di daftar.
    """
    items = await db.pegawai.find(
        scope_query_field_satker(_user), _PROJ).sort("nama", 1).to_list(20000)
    hari_ini = datetime.now(timezone.utc).date().isoformat()
    for it in items:
        it["info_masa"] = info_masa_pegawai(it, hari_ini)
    return {"items": items, "jumlah": len(items)}


@pegawai_router.get("/pegawai/deteksi-identitas")
async def deteksi_identitas_pegawai(nomor: str = "",
                                    _user: dict = Depends(require_user)):
    """Kenali jenis nomor identitas (NIP PNS / NI PPPK / NRP / NIK) dari
    formatnya — utk label & saran otomatis di form (satu sumber logika)."""
    return deteksi_identitas(nomor)


async def _jumlah_aset_per_nip(user) -> dict:
    """Peta NIP → jumlah aset yang dipegang (aset ter-scope satker user)."""
    from shared_utils import scope_query_aset
    from shared_utils import filter_aset_perhitungan
    q = await filter_aset_perhitungan(await scope_query_aset(user, {
        "dihapus": {"$ne": True},
        "pengguna_nip": {"$nin": ["", None]},
    }))
    peta = {}
    async for g in db.assets.aggregate([
            {"$match": q},
            {"$group": {"_id": "$pengguna_nip", "n": {"$sum": 1}}}]):
        nip = str(g["_id"] or "").strip()
        if nip:
            peta[nip] = g["n"]
    return peta


@pegawai_router.get("/pegawai/perlu-serah-terima")
async def daftar_perlu_serah_terima(_user: dict = Depends(require_user)):
    """Pegawai BERISIKO yang masih memegang aset — status keluar/mutasi/
    pensiun/nonaktif/diperbantukan atau kontrak Non-ASN habis/segera habis
    (pola alert "pemegang keluar"): bahan tindak lanjut serah terima BMN via
    modul Penggunaan (BAST pengembalian / mutasi pemegang)."""
    pegawai = await db.pegawai.find(
        scope_query_field_satker(_user),
        {"_id": 0, "id": 1, "nama": 1, "nip": 1, "status": 1,
         "unit_kerja": 1, "tgl_selesai_kontrak": 1}).to_list(20000)
    peta_aset = await _jumlah_aset_per_nip(_user)
    hari_ini = datetime.now(timezone.utc).date().isoformat()
    items = pegawai_perlu_serah_terima(pegawai, peta_aset, hari_ini)
    # lengkapi unit kerja utk tampilan
    unit = {p.get("id"): p.get("unit_kerja") for p in pegawai}
    for it in items:
        it["unit_kerja"] = unit.get(it["id"], "")
    return {"items": items, "jumlah": len(items)}


@pegawai_router.get("/pegawai/{pegawai_id}/aset")
async def aset_dipegang_pegawai(pegawai_id: str,
                                _user: dict = Depends(require_user)):
    """Daftar aset yang tercatat dipegang seorang pegawai (via NIP)."""
    from shared_utils import pastikan_akses_dok_satker, scope_query_aset
    peg = await db.pegawai.find_one({"id": pegawai_id}, _PROJ)
    if not peg:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    await pastikan_akses_dok_satker(_user, peg)
    nip = str(peg.get("nip") or "").strip()
    if not nip:
        return {"pegawai": {"id": pegawai_id, "nama": peg.get("nama"),
                            "nip": ""}, "items": [], "jumlah": 0}
    from shared_utils import filter_aset_perhitungan
    q = await filter_aset_perhitungan(await scope_query_aset(_user, {
        "pengguna_nip": nip, "dihapus": {"$ne": True}}))
    items = await db.assets.find(q, {
        "_id": 0, "id": 1, "asset_code": 1, "NUP": 1, "asset_name": 1,
        "condition": 1, "location": 1, "activity_id": 1, "bast_file_id": 1,
    }).sort("asset_code", 1).to_list(2000)
    return {"pegawai": {"id": pegawai_id, "nama": peg.get("nama"),
                        "nip": nip},
            "items": items, "jumlah": len(items)}


async def _sinkron_pemegang_aset(user, peg, tulis: bool) -> dict:
    """Hitung (dan bila `tulis`, terapkan) penyegaran snapshot pemegang
    (nama/jabatan/melekat-ke) pada aset AKTIF yang dipegang pegawai — match
    NIP, ter-scope satker. Idempoten; hanya menimpa field yang berbeda & data
    master non-kosong. TIDAK menyentuh `bast_file_id` (BAST historis tetap).
    Kembalikan {jumlah_aset, perlu_sinkron, diperbarui, per_field, target}."""
    from pymongo import UpdateOne

    from shared_utils import scope_query_aset
    nip = str((peg or {}).get("nip") or "").strip()
    target = snapshot_pemegang_aset(peg)
    hasil = {"jumlah_aset": 0, "perlu_sinkron": 0, "diperbarui": 0,
             "per_field": {}, "target": target}
    if not nip:
        return hasil
    # Seluruh aset satker (lintas kegiatan, termasuk yang belum sah) yang
    # dipegang orang ini — snapshot pemegang harus terkini di mana pun.
    q = await scope_query_aset(user, {"pengguna_nip": nip,
                                      "dihapus": {"$ne": True}})
    assets = await db.assets.find(q, {
        "_id": 0, "id": 1, "user": 1, "pengguna_jabatan": 1,
        "pengguna_melekat_ke": 1}).to_list(200000)
    hasil["jumlah_aset"] = len(assets)
    now = datetime.now(timezone.utc).isoformat()
    ops = []
    for a in assets:
        beda = beda_snapshot_pemegang(a, target)
        if not beda:
            continue
        hasil["perlu_sinkron"] += 1
        for k in beda:
            hasil["per_field"][k] = hasil["per_field"].get(k, 0) + 1
        if tulis:
            ops.append(UpdateOne({"id": a["id"]},
                                 {"$set": {**beda, "updated_at": now},
                                  "$inc": {"version": 1}}))
    if tulis and ops:
        res = await db.assets.bulk_write(ops, ordered=False)
        hasil["diperbarui"] = res.modified_count or 0
    return hasil


@pegawai_router.get("/pegawai/{pegawai_id}/sinkron-aset/pratinjau")
async def pratinjau_sinkron_aset(pegawai_id: str,
                                 _user: dict = Depends(require_user)):
    """Pratinjau (tanpa menulis): berapa aset dipegang & berapa perlu
    disegarkan datanya bila disinkronkan dengan master pegawai saat ini."""
    peg = await db.pegawai.find_one({"id": pegawai_id}, _PROJ)
    if not peg:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    await pastikan_akses_dok_satker(_user, peg)
    hasil = await _sinkron_pemegang_aset(_user, peg, tulis=False)
    return {"ok": True, **hasil}


@pegawai_router.post("/pegawai/{pegawai_id}/sinkron-aset")
async def sinkron_aset_pemegang(pegawai_id: str,
                                user: dict = Depends(require_admin)):
    """Segarkan data pemegang (nama/jabatan/unit) pada aset AKTIF yang dipegang
    pegawai ini (match NIP) agar DBR/KIR/BAST BERIKUTNYA memakai data terkini —
    mis. setelah kenaikan pangkat / perpindahan unit / perubahan nama. BAST
    yang SUDAH terbit tidak berubah (dokumen historis). Idempoten."""
    peg = await db.pegawai.find_one({"id": pegawai_id}, _PROJ)
    if not peg:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    await pastikan_akses_dok_satker(user, peg)
    if not str(peg.get("nip") or "").strip():
        raise HTTPException(status_code=400,
                            detail="Pegawai belum punya NIP — tidak ada aset tertaut")
    hasil = await _sinkron_pemegang_aset(user, peg, tulis=True)
    if hasil["diperbarui"]:
        await log_audit("sinkron_aset_pemegang", "", pegawai_id,
                        username=user.get("username", "system"),
                        detail=(f"Sinkron data pemegang {peg.get('nama')} ke "
                                f"{hasil['diperbarui']} aset"),
                        kode_satker=str(peg.get("kode_satker") or ""))
    return {"ok": True, **hasil}


@pegawai_router.post("/pegawai")
async def buat_pegawai(payload: PegawaiIn, user: dict = Depends(require_admin)):
    """Tambah pegawai (admin)."""
    doc = _bersih(payload)
    errors = validate_pegawai(doc)
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))
    kode = kode_satker_user(user)
    if await _nip_bentrok(doc["nip"], kode=kode):
        raise HTTPException(status_code=400, detail=f"NIP {doc['nip']} sudah terdaftar")
    now = datetime.now(timezone.utc).isoformat()
    # Admin terikat satker: dipaksa satkernya (isolasi M-SCOPE);
    # super-admin ("" ): boleh mengisi kode_satker eksplisit dari form.
    doc.update({"id": str(uuid.uuid4()), "kode_satker": kode or doc.get("kode_satker", ""),
                "created_at": now, "updated_at": now})
    await db.pegawai.insert_one(dict(doc))
    await log_audit("buat_pegawai", "", doc["id"],
                    username=user.get("username", "system"),
                    detail=f"Tambah pegawai {doc['nama']}",
                    kode_satker=str(doc.get("kode_satker") or ""))
    return {"ok": True, "id": doc["id"]}


@pegawai_router.get("/pegawai/template-impor")
async def template_impor_pegawai(_user: dict = Depends(require_user)):
    """Template CSV impor massal pegawai (header lengkap + 1 baris contoh)."""
    import csv as _csv
    import io

    from pegawai_utils import HEADER_IMPOR
    buf = io.StringIO()
    w = _csv.writer(buf)
    w.writerow(HEADER_IMPOR)
    w.writerow(["198501012010011001", "Budi Santoso", "Laki-laki", "Jakarta",
                "1985-01-01", "PNS", "Penata (III/c)",
                "Analis Pengelolaan BMN", "", "",
                "Pejabat Pelaksana",
                "2022-01-01", "",
                "Sekretariat", "Bagian Umum", "", "", "", "081200000000",
                "budi@instansi.go.id", "", "S1", "",
                "BRI", "1234567890", "", "", "", "", "", "",
                "AKTIF", ""])
    return Response(
        content=buf.getvalue().encode("utf-8-sig"), media_type="text/csv",
        headers={"Content-Disposition":
                 'attachment; filename="template_impor_pegawai.csv"'})


@pegawai_router.get("/pegawai/export-xlsx")
async def export_pegawai_xlsx(_user: dict = Depends(require_user)):
    """Ekspor Master Pegawai ke Excel siap-edit (W7).

    Header = HEADER_IMPOR sehingga file hasil ekspor bisa langsung diedit
    lalu DIIMPOR KEMBALI. Kolom pilihan diberi dropdown (data validation
    dari sheet Referensi), kolom nomor dipaksa format teks (NIP tidak rusak
    jadi float), header beku + berwarna + auto-filter."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    from pegawai_utils import (HEADER_IMPOR, KOLOM_TEKS_EKSPOR,
                               OPSI_DROPDOWN_EKSPOR, baris_ekspor_pegawai)

    items = await db.pegawai.find(
        scope_query_field_satker(_user), _PROJ).sort("nama", 1).to_list(20000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Pegawai"

    biru = "1E40AF"
    header_fill = PatternFill("solid", fgColor=biru)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    zebra_fill = PatternFill("solid", fgColor="EFF6FF")
    tipis = Side(style="thin", color="CBD5E1")
    border = Border(left=tipis, right=tipis, top=tipis, bottom=tipis)

    ws.append(HEADER_IMPOR)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 30

    kolom_teks = {i + 1 for i, h in enumerate(HEADER_IMPOR)
                  if h in KOLOM_TEKS_EKSPOR}
    for ridx, p in enumerate(items, start=2):
        for cidx, val in enumerate(baris_ekspor_pegawai(p), start=1):
            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.border = border
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center")
            if cidx in kolom_teks:
                cell.number_format = "@"
            if ridx % 2 == 0:
                cell.fill = zebra_fill

    # Lebar kolom proporsional isi header/nilai umum
    lebar = {"NIP/NIK/NRP": 22, "Nama Lengkap": 28, "Jabatan": 28,
             "Status Kepegawaian": 24, "Email": 26, "Alamat": 32,
             "Eselon 1": 22, "Eselon 2": 22, "Eselon 3": 22,
             "Eselon 4": 22, "Eselon 5": 22, "Keterangan": 28,
             "Perusahaan Penyedia": 24, "Jenis Kontrak Non-ASN": 30,
             "Kategori Pegawai": 26, "Pangkat/Golongan": 22}
    for i, h in enumerate(HEADER_IMPOR, start=1):
        ws.column_dimensions[get_column_letter(i)].width = lebar.get(h, 15)
    ws.freeze_panes = "C2"   # header + kolom NIP & Nama tetap terlihat
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADER_IMPOR))}1"

    # ── Sheet Referensi: sumber dropdown + panduan nilai sah ──
    ref = wb.create_sheet("Referensi")
    ref["A1"] = ("Daftar nilai sah per kolom — dipakai dropdown di sheet Data "
                 "Pegawai. Boleh menulis nilai lain; impor menormalkan "
                 "sebisanya.")
    ref["A1"].font = Font(bold=True, size=10)
    posisi_ref = {}
    kolom_ref = 1
    for judul, opsi in OPSI_DROPDOWN_EKSPOR.items():
        kol = get_column_letter(kolom_ref)
        sel_judul = ref.cell(row=3, column=kolom_ref, value=judul)
        sel_judul.font = Font(bold=True, color="FFFFFF", size=10)
        sel_judul.fill = header_fill
        for j, v in enumerate(opsi, start=4):
            ref.cell(row=j, column=kolom_ref, value=v)
        posisi_ref[judul] = f"Referensi!${kol}$4:${kol}${3 + len(opsi)}"
        ref.column_dimensions[kol].width = max(
            [len(judul)] + [len(str(v)) for v in opsi]) + 3
        kolom_ref += 1

    # Data validation dropdown — berlaku juga utk 500 baris kosong tambahan
    baris_akhir = max(len(items) + 1, 2) + 500
    for judul, rentang in posisi_ref.items():
        if judul not in HEADER_IMPOR:
            continue
        idx = HEADER_IMPOR.index(judul) + 1
        kol = get_column_letter(idx)
        dv = DataValidation(type="list", formula1=f"={rentang}",
                            allow_blank=True, showDropDown=False)
        dv.error = "Pilih dari daftar atau kosongkan"
        dv.errorTitle = "Nilai di luar daftar"
        dv.errorStyle = "warning"   # tetap boleh nilai bebas (soft)
        ws.add_data_validation(dv)
        dv.add(f"{kol}2:{kol}{baris_akhir}")

    buf = io.BytesIO()
    wb.save(buf)
    kode = kode_satker_user(_user)
    nama_file = f"master_pegawai{('_' + kode) if kode else ''}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition":
                 f'attachment; filename="{nama_file}"'})


@pegawai_router.post("/pegawai/impor")
async def impor_pegawai(file: UploadFile = File(...),
                        user: dict = Depends(require_admin)):
    """Impor massal pegawai dari Excel (.xlsx) / CSV. Tiap baris dinormalkan
    (status kepegawaian & status keberadaan yang beragam dipetakan, NIP
    dibersihkan dari artefak, unit kerja diambil dari Eselon terdalam). Upsert
    per NIP dalam satker; baris tanpa NIP disisipkan baru. Data lapangan yang
    tak baku hanya diberi catatan (best-effort), tidak menggagalkan seluruh
    impor."""
    nama = str(file.filename or "").lower()
    data = await file.read()
    if len(data) > 15 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Berkas maksimal 15MB")

    rows = []
    if nama.endswith((".xlsx", ".xlsm")):
        import io

        from openpyxl import load_workbook
        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(status_code=400,
                                detail="Berkas Excel tidak dapat dibaca")
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h or "").strip() for h in next(it)]
        except StopIteration:
            header = []
        for r in it:
            rows.append({header[i]: r[i]
                         for i in range(min(len(header), len(r)))})
    elif nama.endswith(".csv"):
        import csv as _csv
        import io
        reader = _csv.DictReader(io.StringIO(data.decode("utf-8-sig",
                                                         errors="replace")))
        rows = list(reader)
    else:
        raise HTTPException(status_code=400,
                            detail="Format harus .xlsx atau .csv")

    kode = kode_satker_user(user)
    now = datetime.now(timezone.utc).isoformat()
    dilewati = 0
    catatan = []
    seen_nip = set()
    from pymongo import InsertOne, UpdateOne
    ops = []
    for idx, raw in enumerate(rows, start=2):
        doc, _peringatan = baris_impor_ke_pegawai(raw)
        if not doc["nama"]:
            dilewati += 1
            continue
        # Validasi lunak: abaikan keluhan NIP (data lapangan sering non-baku).
        errors = [e for e in validate_pegawai(doc) if "NIP" not in e]
        if errors:
            dilewati += 1
            if len(catatan) < 200:
                catatan.append(f"Baris {idx} ({doc['nama']}): {'; '.join(errors)}")
            continue
        doc["kode_satker"] = kode
        doc["updated_at"] = now
        nip = doc["nip"]
        if nip and nip not in seen_nip:
            seen_nip.add(nip)
            ops.append(UpdateOne(
                {"nip": nip, "kode_satker": {"$in": [kode, "", None]}},
                {"$set": doc,
                 "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now}},
                upsert=True))
        else:
            doc["id"] = str(uuid.uuid4())
            doc["created_at"] = now
            ops.append(InsertOne(dict(doc)))

    dibuat = diperbarui = 0
    if ops:
        res = await db.pegawai.bulk_write(ops, ordered=False)
        dibuat = (res.upserted_count or 0) + (res.inserted_count or 0)
        diperbarui = res.modified_count or 0
    await log_audit("impor_pegawai", "", "impor",
                    username=user.get("username", "system"),
                    detail=(f"Impor pegawai: {dibuat} baru, {diperbarui} "
                            f"diperbarui, {dilewati} dilewati"),
                    kode_satker=kode_satker_user(user))
    return {"ok": True, "dibaca": len(rows), "dibuat": dibuat,
            "diperbarui": diperbarui, "dilewati": dilewati,
            "catatan": catatan}


@pegawai_router.put("/pegawai/{pegawai_id}")
async def ubah_pegawai(pegawai_id: str, payload: PegawaiIn,
                       user: dict = Depends(require_admin)):
    """Ubah pegawai (admin)."""
    from shared_utils import pastikan_akses_dok_satker
    doc = _bersih(payload)
    errors = validate_pegawai(doc)
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))
    # Isolasi satker: admin terikat tak boleh mengubah pegawai satker lain.
    # Identitas lama ikut dibaca untuk mendeteksi perlunya sinkron aset.
    lama = await db.pegawai.find_one(
        {"id": pegawai_id},
        {"_id": 0, "kode_satker": 1, "nama": 1, "gelar_depan": 1,
         "gelar_belakang": 1, "jabatan": 1, "unit_kerja": 1})
    if not lama:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    await pastikan_akses_dok_satker(user, lama)
    if await _nip_bentrok(doc["nip"], kecuali_id=pegawai_id,
                          kode=kode_satker_user(user)):
        raise HTTPException(status_code=400, detail=f"NIP {doc['nip']} sudah terdaftar")
    # Isolasi M-SCOPE: admin terikat TIDAK bisa memindah pegawai ke satker
    # lain lewat payload; super-admin boleh — kosong = pertahankan yang lama.
    kode_user = kode_satker_user(user)
    if kode_user:
        doc["kode_satker"] = kode_user
    elif not doc.get("kode_satker"):
        doc["kode_satker"] = str(lama.get("kode_satker") or "")
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.pegawai.update_one({"id": pegawai_id}, {"$set": doc})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    await log_audit("ubah_pegawai", "", pegawai_id,
                    username=user.get("username", "system"),
                    detail=f"Ubah pegawai {doc['nama']}",
                    kode_satker=str(doc.get("kode_satker") or ""))
    # Peringatan lunak: status berubah ke non-aktif padahal masih memegang
    # aset → dorong proses serah terima (tidak memblokir penyimpanan).
    peringatan = ""
    if doc["status"] not in ("aktif", "cuti", "tugas_belajar") and doc["nip"]:
        dipegang = await db.assets.count_documents(
            {"pengguna_nip": doc["nip"], "dihapus": {"$ne": True}})
        if dipegang:
            peringatan = (
                f"{doc['nama']} masih tercatat memegang {dipegang} aset — "
                "proses serah terima di modul Penggunaan (BAST pengembalian "
                "atau mutasi pemegang)")
    # Saran sinkron aset: bila identitas (nama/jabatan/unit) berubah & pegawai
    # masih memegang aset, tawarkan penyegaran snapshot pemegang pada aset
    # (DBR/KIR/BAST berikutnya) — BAST historis tidak diubah.
    sinkron = None
    ident_berubah = (
        snapshot_pemegang_aset({**lama, "id": pegawai_id})
        != snapshot_pemegang_aset({**doc, "id": pegawai_id}))
    if doc["nip"] and ident_berubah:
        h = await _sinkron_pemegang_aset(user, {**doc, "id": pegawai_id},
                                         tulis=False)
        if h["perlu_sinkron"]:
            sinkron = {"jumlah_aset": h["jumlah_aset"],
                       "perlu_sinkron": h["perlu_sinkron"],
                       "per_field": h["per_field"]}
    return {"ok": True, "id": pegawai_id, "peringatan": peringatan,
            "sinkron_aset": sinkron}


@pegawai_router.delete("/pegawai/{pegawai_id}")
async def hapus_pegawai(pegawai_id: str, user: dict = Depends(require_admin)):
    """Hapus pegawai (admin). Ditolak bila NIP-nya masih dipakai aset (temuan #34)."""
    from shared_utils import pastikan_akses_dok_satker
    peg = await db.pegawai.find_one({"id": pegawai_id}, {"_id": 0, "nip": 1, "nama": 1, "kode_satker": 1})
    if not peg:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    await pastikan_akses_dok_satker(user, peg)  # isolasi satker
    if peg and str(peg.get("nip") or "").strip():
        dipakai = await db.assets.count_documents(
            {"pengguna_nip": str(peg["nip"]).strip(), "dihapus": {"$ne": True}})
        if dipakai:
            raise HTTPException(
                status_code=409,
                detail=f"Pegawai {peg.get('nama')} (NIP {peg['nip']}) masih tercatat "
                       f"sebagai pengguna pada {dipakai} aset — pindahkan aset atau "
                       f"ubah status pegawai menjadi nonaktif, jangan dihapus.")
    res = await db.pegawai.delete_one({"id": pegawai_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    await log_audit("hapus_pegawai", "", pegawai_id,
                    username=user.get("username", "system"),
                    detail=f"Hapus pegawai {peg.get('nama') or ''}".strip(),
                    kode_satker=str(peg.get("kode_satker") or ""))
    return {"ok": True, "id": pegawai_id}


# ============================================================================
# FOTO PEGAWAI — persegi hasil krop dari frontend (zoom/geser), GridFS.
# Otomatis tercakup backup/restore (ekspor GridFS penuh di modul Backup).
# ============================================================================

_FOTO_MAX = 5 * 1024 * 1024  # 5MB


def _kecilkan_foto_asli(data: bytes, maks_sisi: int = 1600) -> bytes:
    """Kecilkan foto ASLI (sumber krop) agar hemat penyimpanan namun cukup
    tajam untuk diposisikan ulang (keluaran avatar hanya 384px). Datar-kan
    transparansi ke putih & simpan JPEG. Bila Pillow gagal, kembalikan apa
    adanya (best-effort)."""
    try:
        import io as _io

        from PIL import Image
        img = Image.open(_io.BytesIO(data))
        if img.mode in ("RGBA", "LA", "P"):
            latar = Image.new("RGB", img.size, (255, 255, 255))
            img = img.convert("RGBA")
            latar.paste(img, mask=img.split()[-1])
            img = latar
        else:
            img = img.convert("RGB")
        w, h = img.size
        if max(w, h) > maks_sisi:
            skala = maks_sisi / float(max(w, h))
            img = img.resize((max(1, int(w * skala)), max(1, int(h * skala))),
                             Image.LANCZOS)
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=85, optimize=True)
        return buf.getvalue()
    except Exception:
        return data


@pegawai_router.post("/pegawai/{pegawai_id}/foto")
async def upload_foto_pegawai(pegawai_id: str, file: UploadFile = File(...),
                              file_asli: UploadFile = File(None),
                              krop: str = Form(""),
                              admin: dict = Depends(require_admin)):
    """Unggah foto pegawai (persegi hasil krop di frontend); ganti = hapus lama.

    Opsional simpan **foto ASLI** (`file_asli`, dikecilkan) + **parameter krop**
    (`krop` JSON: zoom/posisi) agar posisi foto dapat DIATUR ULANG tanpa memilih
    berkas lagi. `file_asli` kosong = pertahankan asli lama (reposisi saja)."""
    peg = await db.pegawai.find_one(
        {"id": pegawai_id},
        {"_id": 0, "id": 1, "foto_file_id": 1, "foto_asli_file_id": 1,
         "nama": 1, "kode_satker": 1})
    if not peg:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    if (file.content_type or "") not in ("image/jpeg", "image/png", "image/webp"):
        raise HTTPException(status_code=400, detail="Format harus JPG/PNG/WebP")
    data = await file.read()
    if len(data) > _FOTO_MAX:
        raise HTTPException(status_code=400, detail="Ukuran foto maksimal 5MB")
    from bson import ObjectId
    fid = await fs_bucket.upload_from_stream(
        f"pegawai_{pegawai_id}.jpg", data,
        metadata={"jenis": "foto_pegawai", "pegawai_id": pegawai_id,
                  "content_type": file.content_type})
    now = datetime.now(timezone.utc).isoformat()
    perubahan = {"foto_file_id": str(fid), "updated_at": now}

    # Foto asli (sumber krop) — hanya diganti bila berkas asli baru dikirim.
    if file_asli is not None:
        asli_raw = await file_asli.read()
        if len(asli_raw) > _FOTO_MAX:
            raise HTTPException(status_code=400, detail="Ukuran foto asli maksimal 5MB")
        asli = _kecilkan_foto_asli(asli_raw)
        aid = await fs_bucket.upload_from_stream(
            f"pegawai_{pegawai_id}_asli.jpg", asli,
            metadata={"jenis": "foto_pegawai_asli", "pegawai_id": pegawai_id,
                      "content_type": "image/jpeg"})
        lama_asli = peg.get("foto_asli_file_id")
        if lama_asli:
            try:
                await fs_bucket.delete(ObjectId(lama_asli))
            except Exception:
                pass
        perubahan["foto_asli_file_id"] = str(aid)

    # Parameter krop (untuk seed dialog reposisi) — best-effort parse.
    if str(krop or "").strip():
        import json
        try:
            perubahan["foto_krop"] = json.loads(krop)
        except (ValueError, TypeError):
            pass

    lama = peg.get("foto_file_id")
    if lama:
        try:
            await fs_bucket.delete(ObjectId(lama))
        except Exception:
            pass
    await db.pegawai.update_one({"id": pegawai_id}, {"$set": perubahan})
    await log_audit("foto_pegawai", "", username=admin.get("username", "system"),
                    detail=f"Foto pegawai {peg.get('nama')} diperbarui",
                    kode_satker=str(peg.get("kode_satker") or ""))
    return {"ok": True, "foto_file_id": str(fid),
            "foto_asli_file_id": perubahan.get("foto_asli_file_id",
                                               peg.get("foto_asli_file_id") or "")}


@pegawai_router.get("/pegawai/{pegawai_id}/foto-asli")
async def get_foto_asli_pegawai(pegawai_id: str,
                                _user: dict = Depends(require_user_or_query_token)):
    """Stream foto ASLI (sumber krop) — dipakai dialog "Atur Ulang Posisi".
    404 bila pegawai/foto era lama yang belum menyimpan berkas asli."""
    peg = await db.pegawai.find_one({"id": pegawai_id},
                                    {"_id": 0, "foto_asli_file_id": 1})
    if not peg or not peg.get("foto_asli_file_id"):
        raise HTTPException(status_code=404, detail="Foto asli tidak tersedia")
    from bson import ObjectId
    try:
        stream = await fs_bucket.open_download_stream(
            ObjectId(peg["foto_asli_file_id"]))
        data = await stream.read()
    except Exception:
        raise HTTPException(status_code=404,
                            detail="Foto asli tidak ditemukan di penyimpanan")
    ct = (stream.metadata or {}).get("content_type") or "image/jpeg"
    return Response(content=data, media_type=ct,
                    headers={"Cache-Control": "private, max-age=86400"})


@pegawai_router.get("/pegawai/{pegawai_id}/foto")
async def get_foto_pegawai(pegawai_id: str,
                           _user: dict = Depends(require_user_or_query_token)):
    """Stream foto pegawai (dipakai avatar row + form; cache per file_id)."""
    peg = await db.pegawai.find_one({"id": pegawai_id}, {"_id": 0, "foto_file_id": 1})
    if not peg or not peg.get("foto_file_id"):
        raise HTTPException(status_code=404, detail="Foto tidak ada")
    from bson import ObjectId
    try:
        stream = await fs_bucket.open_download_stream(ObjectId(peg["foto_file_id"]))
        data = await stream.read()
    except Exception:
        raise HTTPException(status_code=404, detail="Foto tidak ditemukan di penyimpanan")
    ct = (stream.metadata or {}).get("content_type") or "image/jpeg"
    return Response(content=data, media_type=ct,
                    headers={"Cache-Control": "private, max-age=86400"})


@pegawai_router.delete("/pegawai/{pegawai_id}/foto")
async def hapus_foto_pegawai(pegawai_id: str, admin: dict = Depends(require_admin)):
    """Hapus foto pegawai (GridFS + referensi) — termasuk foto asli & krop."""
    peg = await db.pegawai.find_one(
        {"id": pegawai_id},
        {"_id": 0, "foto_file_id": 1, "foto_asli_file_id": 1})
    if not peg or not peg.get("foto_file_id"):
        raise HTTPException(status_code=404, detail="Foto tidak ada")
    from bson import ObjectId
    for fid in (peg.get("foto_file_id"), peg.get("foto_asli_file_id")):
        if fid:
            try:
                await fs_bucket.delete(ObjectId(fid))
            except Exception:
                pass
    now = datetime.now(timezone.utc).isoformat()
    await db.pegawai.update_one(
        {"id": pegawai_id},
        {"$unset": {"foto_file_id": "", "foto_asli_file_id": "", "foto_krop": ""},
         "$set": {"updated_at": now}})
    return {"ok": True}


# ============================================================================
# KARTU PEGAWAI (UID e-KTP / kartu NFC) — pendaftaran & identifikasi tap
#
# UID kartu TIDAK PERNAH disimpan/di-log mentah — hanya HMAC-SHA256 berkunci
# rahasia server (kartu_utils). Multi-hash per kartu menoleransi perbedaan
# format keluaran reader (hex MSB/LSB, desimal). Field ikut dokumen pegawai
# sehingga otomatis ter-backup & selamat reset (RESET_KEEP pegawai).
# Batas peran: UID = identifikasi cepat/kenyamanan (pengganti mengetik),
# BUKAN verifikasi keaslian KTP dan BUKAN pengganti TTD elektronik.
# ============================================================================

_KARTU_PROYEKSI_PUBLIK = {
    "_id": 0, "id": 1, "nama": 1, "nip": 1, "jabatan": 1, "unit_kerja": 1,
    "email": 1, "status": 1, "status_kepegawaian": 1, "foto_file_id": 1,
    "kartu_label": 1, "kode_satker": 1,
}


class KartuIn(BaseModel):
    uid: str = Field(min_length=1, max_length=64)


@pegawai_router.post("/pegawai/kartu/identifikasi")
@limiter.limit("30/minute")
async def identifikasi_kartu(request: Request, payload: KartuIn,
                             user: dict = Depends(require_user)):
    """Tap kartu → pegawai pemiliknya (identifikasi cepat lintas modul).

    Dipakai komponen Tap Kartu di form aset, BAST, TTD elektronik, dsb.
    Ber-rate-limit + terautentikasi; kartu tak dikenal dicatat ke audit
    (sinyal keamanan) tanpa UID mentah — hanya prefiks hash.
    """
    err = valid_uid(payload.uid)
    if err:
        raise HTTPException(status_code=400, detail=err)
    hashes = hash_kandidat(payload.uid)
    q = scope_query_field_satker(user, {"kartu_uid_hashes": {"$in": hashes}})
    peg = await db.pegawai.find_one(q, _KARTU_PROYEKSI_PUBLIK)
    if not peg:
        await log_audit("kartu_tak_dikenal", "", "",
                        username=user.get("username", "system"),
                        detail=f"Tap kartu tidak dikenal (hash {hashes[0][:8]}…)",
                        kode_satker=kode_satker_user(user))
        raise HTTPException(status_code=404,
                            detail="Kartu tidak dikenal — daftarkan dulu di Master Pegawai")
    return {"pegawai": peg}


@pegawai_router.post("/pegawai/{pegawai_id}/kartu")
@limiter.limit("10/minute")
async def daftar_kartu_pegawai(request: Request, pegawai_id: str,
                               payload: KartuIn,
                               admin: dict = Depends(require_admin)):
    """Daftarkan (atau ganti) kartu e-KTP/NFC seorang pegawai — cukup tap.

    Satu kartu hanya boleh terikat SATU pegawai (cek silang semua hash
    kandidat). Pendaftaran ulang menimpa kartu lama pegawai tsb.
    """
    err = valid_uid(payload.uid)
    if err:
        raise HTTPException(status_code=400, detail=err)
    peg = await db.pegawai.find_one({"id": pegawai_id}, _PROJ)
    if not peg:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    await pastikan_akses_dok_satker(admin, peg)
    hashes = hash_kandidat(payload.uid)
    lain = await db.pegawai.find_one(
        {"kartu_uid_hashes": {"$in": hashes}, "id": {"$ne": pegawai_id}},
        {"_id": 0, "nama": 1, "kode_satker": 1})
    if lain:
        # Nama pegawai konflik hanya disebut bila SE-SATKER dengan admin —
        # jangan bocorkan nama pegawai satker lain lewat probing UID.
        kode_admin = kode_satker_user(admin)
        se_satker = (not kode_admin
                     or str(lain.get("kode_satker") or "").strip() in ("", kode_admin))
        pesan = (f"Kartu ini sudah terdaftar pada pegawai lain "
                 f"({lain.get('nama') or '-'}). Lepas dulu dari pegawai tersebut."
                 if se_satker else
                 "Kartu ini sudah terdaftar pada pegawai satuan kerja lain — "
                 "hubungi administrator pusat.")
        raise HTTPException(status_code=409, detail=pesan)
    now = datetime.now(timezone.utc).isoformat()
    label = label_kartu(payload.uid)
    try:
        await db.pegawai.update_one({"id": pegawai_id}, {"$set": {
            "kartu_uid_hashes": hashes,
            "kartu_label": label,
            "kartu_terdaftar_pada": now,
            "kartu_terdaftar_oleh": admin.get("username", "system"),
            "updated_at": now,
        }})
    except Exception:
        # Index unik kartu_uid_hashes (bila berhasil terpasang) menutup
        # celah balapan dua admin mendaftarkan kartu sama bersamaan.
        raise HTTPException(status_code=409,
                            detail="Kartu baru saja terdaftar pada pegawai lain — muat ulang lalu coba lagi.")
    await log_audit("daftar_kartu", "", pegawai_id,
                    username=admin.get("username", "system"),
                    detail=f"Daftarkan kartu {label} untuk {peg.get('nama') or '-'}",
                    kode_satker=str(peg.get("kode_satker") or ""))
    return {"ok": True, "kartu_label": label, "kartu_terdaftar_pada": now}


@pegawai_router.delete("/pegawai/{pegawai_id}/kartu")
@limiter.limit("10/minute")
async def lepas_kartu_pegawai(request: Request, pegawai_id: str,
                              admin: dict = Depends(require_admin)):
    """Lepas kartu dari pegawai (kartu hilang / ganti KTP / pegawai keluar)."""
    peg = await db.pegawai.find_one(
        {"id": pegawai_id}, {"_id": 0, "nama": 1, "kartu_label": 1,
                             "kartu_uid_hashes": 1, "kode_satker": 1})
    if not peg or not peg.get("kartu_uid_hashes"):
        raise HTTPException(status_code=404, detail="Pegawai belum punya kartu terdaftar")
    await pastikan_akses_dok_satker(admin, peg)
    now = datetime.now(timezone.utc).isoformat()
    await db.pegawai.update_one({"id": pegawai_id}, {
        "$unset": {"kartu_uid_hashes": "", "kartu_label": "",
                   "kartu_terdaftar_pada": "", "kartu_terdaftar_oleh": ""},
        "$set": {"updated_at": now}})
    await log_audit("lepas_kartu", "", pegawai_id,
                    username=admin.get("username", "system"),
                    detail=f"Lepas kartu {peg.get('kartu_label') or ''} dari {peg.get('nama') or '-'}",
                    kode_satker=str(peg.get("kode_satker") or ""))
    return {"ok": True}
