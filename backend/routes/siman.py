"""SINKRONISASI SIMAN V2 — impor manual ekspor "Master Aset" sebagai kanal
pembaruan berkala (pengganti API yang belum tersedia untuk satker).

Alur: unggah XLSX hasil ekspor SIMAN V2 → tiap baris dicocokkan ke aset
AMAN (Kode Register dulu, lalu Kode Barang+NUP) → perbedaan field kunci
tercatat pada subdoc `siman` aset (tanda di halaman aset) → pengguna
meninjau lalu "terapkan nilai SIMAN" per aset (SIMAN = data valid).
Riwayat impor tersimpan di register `siman_imports`.

Logika banding murni di siman_utils.py (teruji unit).
"""
import io
import uuid
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from pydantic import BaseModel

from auth_utils import require_admin, require_user, require_writer
from db import db
from report_filters import active_asset_filter
from shared_utils import (limiter, log_audit, kode_satker_user,
                          pastikan_akses_aset, pastikan_akses_dok_satker,
                          pastikan_akses_kegiatan_id, scope_query_aset,
                          scope_query_field_satker)
from siman_utils import (
    FIELD_TERAPKAN, banding_aset, deteksi_header, kunci_aset, nilai_terapkan,
    parse_baris, referensi_siman, ringkas_baris_belum_tercatat, ringkas_import,
    validasi_satker,
)

siman_router = APIRouter()

_MAKS_UKURAN_FILE = 25 * 1024 * 1024  # 25MB — jauh di atas ekspor satker wajar
_MAKS_BELUM_TERCATAT = 5000  # baris ringkas yang disimpan utk CSV/buat draft

_PROJ_ASET = {
    "_id": 0, "id": 1, "asset_code": 1, "NUP": 1, "asset_name": 1,
    "category": 1, "brand": 1, "model": 1, "condition": 1, "status": 1,
    "purchase_price": 1, "purchase_date": 1, "user": 1, "kode_register": 1,
    "activity_id": 1, "siman": 1,
}
_MAKS_CONTOH = 200  # batas daftar contoh yang disimpan di register impor


class TerapkanIn(BaseModel):
    fields: Optional[List[str]] = None  # None = terapkan semua selisih


@siman_router.post("/siman/import")
@limiter.limit("6/minute")
async def import_siman(request: Request, file: UploadFile = File(...),
                       tandai_tidak_ditemukan: bool = False,
                       user: dict = Depends(require_admin)):
    """Impor ekspor SIMAN V2 (XLSX "Master Aset") dan tandai selisih per aset.

    `tandai_tidak_ditemukan=true` juga menandai aset AMAN yang TIDAK ada di
    file (pakai hanya bila file memuat SELURUH aset satker — ekspor penuh,
    bukan potongan).
    """
    nama_file = file.filename or ""
    if nama_file.lower().endswith(".xls") and not nama_file.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=(
            "Format .xls lama tidak didukung — buka file lalu simpan ulang "
            "sebagai .xlsx (Excel Workbook), atau unduh ulang ekspor dari SIMAN V2"))
    if not nama_file.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File harus Excel (.xlsx) hasil ekspor SIMAN V2")
    isi = await file.read()
    if len(isi) > _MAKS_UKURAN_FILE:
        raise HTTPException(status_code=400, detail=(
            f"File terlalu besar ({len(isi) // (1024 * 1024)}MB — maks 25MB). "
            "Ekspor per jenis BMN dari SIMAN lalu unggah bertahap"))
    if not isi:
        raise HTTPException(status_code=400, detail=(
            "File kosong terkirim — koneksi kemungkinan terputus saat mengunggah; coba lagi"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(isi), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail=(
            "File Excel tidak dapat dibaca — pastikan file ASLI hasil ekspor "
            "SIMAN V2 (bukan hasil ganti nama ekstensi) dan tidak rusak saat diunduh"))

    # Cari sheet + baris header di SEMUA sheet: prioritas "Master Aset", lalu
    # sheet lain — ekspor SIMAN kadang mengganti nama sheet / menambah kop di
    # baris awal. Metadata dimensi ekspor SIMAN kerap salah (mengaku 1 baris)
    # sehingga WAJIB reset_dimensions sebelum membaca (terbukti pada file asli).
    daftar_sheet = list(wb.sheetnames)
    urutan_sheet = sorted(daftar_sheet, key=lambda n: (n != "Master Aset", n))
    peta_header, baris_data, sheet_dipakai = None, [], None
    for nama_sheet in urutan_sheet:
        ws = wb[nama_sheet]
        ws.reset_dimensions()
        rows = ws.iter_rows(values_only=True)
        awal = [next(rows, None) for _ in range(25)]
        hasil_deteksi = deteksi_header([r for r in awal if r is not None])
        if hasil_deteksi is None:
            continue
        idx_header, peta_header = hasil_deteksi
        sheet_dipakai = nama_sheet
        for row in awal[idx_header + 1:]:
            if row is None:
                continue
            b = parse_baris(row, peta_header)
            if b:
                baris_data.append(b)
        for row in rows:  # lanjutan setelah 25 baris awal
            b = parse_baris(row, peta_header)
            if b:
                baris_data.append(b)
        break
    wb.close()
    if peta_header is None:
        raise HTTPException(status_code=400, detail=(
            "Header tidak dikenali di semua sheet — pastikan file adalah ekspor "
            "SIMAN V2 'Master Aset' dengan kolom 'Kode Barang' dan 'NUP' "
            f"(sheet pada file: {', '.join(daftar_sheet[:5])})"))
    if not baris_data:
        raise HTTPException(status_code=400, detail=(
            f"Tidak ada baris aset pada sheet '{sheet_dipakai}' — file mungkin "
            "ekspor kosong; periksa filter saat mengekspor dari SIMAN V2"))

    # Peta pencocokan: kode register (paling stabil) & kode+NUP.
    per_register, per_kunci, duplikat_kunci = {}, {}, 0
    for b in baris_data:
        if b["kode_register"]:
            per_register.setdefault(b["kode_register"], b)
        k = kunci_aset(b["kode_barang"], b["nup"])
        if k in per_kunci:
            duplikat_kunci += 1
        else:
            per_kunci[k] = b

    # ISOLASI SATKER: impor hanya menyentuh aset satker si pengunggah
    # (super-admin lintas-satker). Cegah penulisan subdoc siman / penandaan
    # "tidak ditemukan" pada aset satker lain.
    assets = await db.assets.find(
        await scope_query_aset(user, active_asset_filter()),
        _PROJ_ASET).to_list(500000)

    now = datetime.now(timezone.utc).isoformat()
    import_id = str(uuid.uuid4())
    hasil, terpakai = [], set()
    for a in assets:
        reg = str(a.get("kode_register") or "").strip()
        b = per_register.get(reg) if reg else None
        if b is None:
            b = per_kunci.get(kunci_aset(a.get("asset_code"), a.get("NUP")))
        if b is None:
            continue
        terpakai.add(id(b))
        selisih = banding_aset(a, b)
        # Sinyal reklasifikasi (riset G7 §5): kode_register cocok tetapi
        # kodefikasi/NUP beda → aset direklasifikasi di SIMAN, bukan sekadar
        # selisih field — UI menampilkannya berbeda.
        from mutasi_bmn_utils import deteksi_reklasifikasi_siman
        reklas = deteksi_reklasifikasi_siman(a, b)
        hasil.append({"aset": a, "baris": b, "selisih": selisih,
                      "reklasifikasi": reklas})

    aset_cocok_id = {r["aset"]["id"] for r in hasil}
    aman_tanpa_siman = [a for a in assets if a["id"] not in aset_cocok_id]
    siman_tanpa_aset = [b for b in baris_data if id(b) not in terpakai]

    # Tulis subdoc `siman` per aset yang cocok (bulk, $inc version).
    from pymongo import UpdateOne
    ops = []
    register_diadopsi = 0
    for r in hasil:
        subdoc = {
            "status": "selisih" if r["selisih"] else "cocok",
            "selisih": r["selisih"],
            "reklasifikasi": r.get("reklasifikasi") or {},
            "referensi": referensi_siman(r["baris"]),
            "kode_register": r["baris"]["kode_register"],
            "import_id": import_id,
            "diperiksa_pada": now,
        }
        set_doc = {"siman": subdoc, "updated_at": now}
        # Adopsi ID register SIMAN bila AMAN belum punya — memperkuat
        # pencocokan impor berikutnya (tahan reklasifikasi kode barang).
        if r["baris"]["kode_register"] and not str(r["aset"].get("kode_register") or "").strip():
            set_doc["kode_register"] = r["baris"]["kode_register"]
            register_diadopsi += 1
        ops.append(UpdateOne({"id": r["aset"]["id"]},
                             {"$set": set_doc, "$inc": {"version": 1}}))
    if tandai_tidak_ditemukan:
        for a in aman_tanpa_siman:
            ops.append(UpdateOne(
                {"id": a["id"]},
                {"$set": {"siman": {"status": "tidak_di_siman", "selisih": [],
                                    "import_id": import_id,
                                    "diperiksa_pada": now},
                          "updated_at": now},
                 "$inc": {"version": 1}}))
    if ops:
        await db.assets.bulk_write(ops, ordered=False)

    # "SIMAN menang" (pilihan pemilik): tiap impor langsung memperbarui
    # referensi masa manfaat kelompok terkait (dipakai Penilaian). Kolom
    # "Umur Aset" SIMAN V2 = SISA masa manfaat dalam SEMESTER (bukan tahun) —
    # masa manfaat diderivasi dari identitas garis lurus perolehan/nilai buku
    # (lihat masa_manfaat_dari_siman). Terkumpul sedikit demi sedikit dari
    # data lapangan → tak bergantung revisi KMK. Ditandai sumber="siman".
    from penilaian_utils import masa_manfaat_dari_siman
    mm_teramati = masa_manfaat_dari_siman(baris_data)
    mm_ops = [
        UpdateOne(
            {"kode": kelompok},
            {"$set": {"kode": kelompok, "tahun": int(info["tahun"]),
                      "sumber": "siman", "observasi": int(info["observasi"]),
                      "updated_at": now, "updated_by": "SIMAN import"},
             "$setOnInsert": {"created_at": now}},
            upsert=True,
        )
        for kelompok, info in mm_teramati.items()
    ]
    if mm_ops:
        await db.masa_manfaat.bulk_write(mm_ops, ordered=False)

    # Validasi satker: kode satker pada FILE vs satker terdaftar di AMAN
    # (master satker + kop global + KEGIATAN inventarisasi — kegiatan kini
    # membawa kode_satker_lengkap ±20 digit versi SIMAN V2, sementara AMAN
    # memakai 6 digit) — file milik satker lain terdeteksi dini, dan file
    # satker sendiri tidak lagi salah diperingatkan.
    # ISOLASI SATKER (REVIEW-9 R15): daftar pembanding dibatasi satker
    # PEMANGGIL. Tanpa ini daftar memuat kode SELURUH satker terdaftar, lalu
    # pesan peringatan di bawah mencetaknya kembali ke pengguna — jadi cukup
    # mengunggah file asal untuk memanen daftar kode satker se-instansi.
    from shared_utils import scope_query_field_satker as _sqfs
    kode_terdaftar = set()
    async for s in db.satker.find(
            _sqfs(user, {}), {"_id": 0, "kode_satker": 1, "kode_satker_lengkap": 1}):
        kode_terdaftar.add(s.get("kode_satker_lengkap") or "")
        kode_terdaftar.add(s.get("kode_satker") or "")
    async for keg in db.inventory_activities.find(
            _sqfs(user, {}), {"_id": 0, "kode_satker": 1, "kode_satker_lengkap": 1}):
        kode_terdaftar.add(keg.get("kode_satker_lengkap") or "")
        kode_terdaftar.add(keg.get("kode_satker") or "")
    setelan = await db.report_settings.find_one(
        {"type": "global"}, {"_id": 0, "kode_satker_lengkap": 1}) or {}
    kode_terdaftar.add(setelan.get("kode_satker_lengkap") or "")
    cek_satker = validasi_satker(
        {b.get("kode_satker") for b in baris_data}, kode_terdaftar)
    peringatan = []
    if not cek_satker["cocok"]:
        peringatan.append(
            "Kode satker pada file ({}) BERBEDA dengan satker terdaftar di AMAN "
            "({}) — pastikan file ekspor milik satker Anda".format(
                ", ".join(cek_satker["kode_file"]) or "-",
                ", ".join(cek_satker["kode_terdaftar"]) or "-"))
    if duplikat_kunci:
        peringatan.append(
            f"{duplikat_kunci} baris duplikat kode+NUP pada file dilewati "
            "(baris pertama yang dipakai)")

    ringkasan = ringkas_import(
        [{"selisih": r["selisih"]} for r in hasil], siman_tanpa_aset, aman_tanpa_siman)
    register = {
        "id": import_id,
        "kode_satker": kode_satker_user(user),
        "filename": nama_file,
        "waktu": now,
        "oleh": user.get("username", "system"),
        "sheet": sheet_dipakai,
        "total_baris": len(baris_data),
        "duplikat_kunci": duplikat_kunci,
        "register_diadopsi": register_diadopsi,
        "masa_manfaat_diperbarui": len(mm_teramati),
        "tandai_tidak_ditemukan": bool(tandai_tidak_ditemukan),
        "ringkasan": ringkasan,
        "peringatan": peringatan,
        "validasi_satker": cek_satker,
        # Baris SIMAN yang belum tercatat di AMAN — bahan CSV & buat draft.
        "baris_belum_tercatat": [
            ringkas_baris_belum_tercatat(b)
            for b in siman_tanpa_aset[:_MAKS_BELUM_TERCATAT]],
        "satker": sorted({b["nama_satker"] for b in baris_data if b.get("nama_satker")})[:10],
        # Contoh (dibatasi) untuk ditinjau di UI — bukan data lengkap.
        "contoh_siman_tanpa_aset": [
            {"kode_barang": b["kode_barang"], "nup": b["nup"],
             "nama_barang": b["nama_barang"], "nilai_perolehan": b["nilai_perolehan"]}
            for b in siman_tanpa_aset[:_MAKS_CONTOH]],
        "contoh_aman_tanpa_siman": [
            {"id": a["id"], "asset_code": a.get("asset_code"), "NUP": a.get("NUP"),
             "asset_name": a.get("asset_name")}
            for a in aman_tanpa_siman[:_MAKS_CONTOH]],
    }
    await db.siman_imports.insert_one({**register})
    await log_audit("import_siman", "", username=user.get("username", "system"),
                    detail=(f"Impor SIMAN V2 '{nama_file}': {len(baris_data)} baris, "
                            f"{ringkasan['cocok']} cocok, {ringkasan['selisih']} selisih, "
                            f"{ringkasan['siman_tanpa_aset']} belum tercatat di AMAN"))
    # Respons tanpa daftar panjang (CSV/draft mengambilnya dari register).
    return {k: v for k, v in register.items() if k != "baris_belum_tercatat"}


async def _rekap_per_kegiatan(user: dict) -> list:
    """Rekap status sinkron SIMAN DIPECAH per kegiatan inventarisasi.

    KENAPA PERLU. Angka global ("12 selisih") tak bisa ditindaklanjuti begitu
    satker punya lebih dari satu kegiatan inventarisasi: operator tahu ada
    selisih, tetapi tidak tahu kegiatan mana yang harus dibuka, siapa yang
    menanganinya, atau apakah kegiatan yang BARU saja ia kerjakan sudah bersih.
    Laporan lapangan menyebutnya persis begitu — "informasi per kegiatannya
    tidak diinformasikan sehingga bingung".

    Satu pipeline agregasi, bukan N kueri per kegiatan: jumlah kegiatan tumbuh
    tiap tahun anggaran dan panel ini dimuat pada tiap kunjungan halaman.
    """
    q = await scope_query_aset(user, active_asset_filter({"siman": {"$exists": True}}))
    baris = await db.assets.aggregate([
        {"$match": q},
        {"$group": {"_id": {"kegiatan": "$activity_id", "status": "$siman.status"},
                    "n": {"$sum": 1}}},
    ]).to_list(5000)

    per = {}
    for b in baris:
        kunci = (b.get("_id") or {}).get("kegiatan") or ""
        status = (b.get("_id") or {}).get("status") or "belum_dicek"
        slot = per.setdefault(kunci, {"activity_id": kunci, "nama_kegiatan": "",
                                      "cocok": 0, "selisih": 0, "tidak_di_siman": 0})
        if status in slot:
            slot[status] += int(b.get("n") or 0)

    # Nama kegiatan diambil sekali untuk SELURUH id sekaligus.
    ids = [k for k in per if k]
    if ids:
        async for a in db.inventory_activities.find(
                {"id": {"$in": ids}}, {"_id": 0, "id": 1, "nama_kegiatan": 1}):
            if a.get("id") in per:
                per[a["id"]]["nama_kegiatan"] = a.get("nama_kegiatan") or ""

    hasil = list(per.values())
    for h in hasil:
        h["total"] = h["cocok"] + h["selisih"] + h["tidak_di_siman"]
        if not h["nama_kegiatan"]:
            # Aset tanpa activity_id, atau kegiatannya sudah dihapus. Jangan
            # disembunyikan: justru baris inilah yang paling perlu dilihat.
            h["nama_kegiatan"] = ("(tanpa kegiatan)" if not h["activity_id"]
                                  else "(kegiatan tak ditemukan)")
    # Yang paling perlu ditindaklanjuti di atas.
    hasil.sort(key=lambda h: (-h["selisih"], -h["tidak_di_siman"], h["nama_kegiatan"]))
    return hasil


@siman_router.get("/siman/ringkasan")
async def ringkasan_siman(_user: dict = Depends(require_user)):
    """Status sinkronisasi terkini + riwayat impor (untuk panel UI)."""
    selisih = await db.assets.count_documents(
        await scope_query_aset(_user, active_asset_filter({"siman.status": "selisih"})))
    cocok = await db.assets.count_documents(
        await scope_query_aset(_user, active_asset_filter({"siman.status": "cocok"})))
    tidak_di_siman = await db.assets.count_documents(
        await scope_query_aset(_user, active_asset_filter({"siman.status": "tidak_di_siman"})))
    belum_dicek = await db.assets.count_documents(
        await scope_query_aset(_user, active_asset_filter({"siman": {"$exists": False}})))
    riwayat = await db.siman_imports.find(
        scope_query_field_satker(_user),
        {"_id": 0, "contoh_siman_tanpa_aset": 0, "contoh_aman_tanpa_siman": 0,
         "baris_belum_tercatat": 0},
    ).sort("waktu", -1).limit(10).to_list(10)
    return {
        "selisih": selisih, "cocok": cocok,
        "tidak_di_siman": tidak_di_siman, "belum_dicek": belum_dicek,
        "per_kegiatan": await _rekap_per_kegiatan(_user),
        "riwayat": riwayat,
        "import_terakhir": riwayat[0] if riwayat else None,
    }


@siman_router.get("/siman/import/{import_id}")
async def detail_import_siman(import_id: str, _user: dict = Depends(require_user)):
    """Detail satu impor termasuk daftar contoh yang tidak cocok."""
    reg = await db.siman_imports.find_one(
        {"id": import_id}, {"_id": 0, "baris_belum_tercatat": 0})
    if not reg:
        raise HTTPException(status_code=404, detail="Riwayat impor tidak ditemukan")
    await pastikan_akses_dok_satker(_user, reg)
    return reg


@siman_router.get("/siman/import/{import_id}/belum-tercatat.csv")
async def csv_belum_tercatat(import_id: str, _user: dict = Depends(require_user)):
    """Unduh CSV baris SIMAN yang belum tercatat di AMAN (bahan tindak lanjut)."""
    import csv

    from fastapi.responses import StreamingResponse

    reg = await db.siman_imports.find_one(
        {"id": import_id},
        {"_id": 0, "baris_belum_tercatat": 1, "filename": 1, "kode_satker": 1})
    if not reg:
        raise HTTPException(status_code=404, detail="Riwayat impor tidak ditemukan")
    await pastikan_akses_dok_satker(_user, reg)
    baris = reg.get("baris_belum_tercatat") or []
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Kode Barang", "NUP", "Nama Barang", "Merk", "Tipe", "Kondisi",
                "Nilai Perolehan", "Tanggal Perolehan", "Kode Register SIMAN"])
    for b in baris:
        w.writerow([b.get("kode_barang", ""), b.get("nup", ""),
                    b.get("nama_barang", ""), b.get("merk", ""),
                    b.get("tipe", ""), b.get("kondisi", ""),
                    b.get("nilai_perolehan", 0), b.get("tanggal_perolehan", ""),
                    b.get("kode_register", "")])
    buf.seek(0)
    return StreamingResponse(
        iter([("\ufeff" + buf.getvalue()).encode("utf-8")]), media_type="text/csv",
        headers={"Content-Disposition":
                 'attachment; filename="siman_belum_tercatat.csv"'})


class BuatDraftSimanIn(BaseModel):
    activity_id: str
    maks: int = 500  # batas aman satu panggilan; ulangi bila lebih


@siman_router.post("/siman/import/{import_id}/buat-draft")
@limiter.limit("3/minute")
async def buat_draft_dari_siman(request: Request, import_id: str,
                                payload: BuatDraftSimanIn,
                                user: dict = Depends(require_admin)):
    """Buat aset DRAFT dari baris SIMAN yang belum tercatat di AMAN.

    Data SIMAN langsung menjadi modal awal aset (kode, NUP, nama, merk, tipe,
    kondisi, nilai, tanggal, kode register) — petugas tinggal melengkapi foto
    dan lokasi di lapangan. Jalur create standar dipakai (kunci kegiatan,
    keunikan kode+NUP, registry, audit tetap berlaku); baris yang sudah punya
    aset (impor ulang) otomatis dilewati.
    """
    from models import AssetCreate
    from routes.assets import buat_aset_draft

    reg = await db.siman_imports.find_one(
        {"id": import_id},
        {"_id": 0, "baris_belum_tercatat": 1, "filename": 1, "kode_satker": 1})
    if not reg:
        raise HTTPException(status_code=404, detail="Riwayat impor tidak ditemukan")
    await pastikan_akses_dok_satker(user, reg)
    baris = reg.get("baris_belum_tercatat") or []
    if not baris:
        raise HTTPException(status_code=400, detail=(
            "Impor ini tidak menyimpan baris belum tercatat — jalankan impor "
            "ulang (fitur ini berlaku untuk impor terbaru)"))
    act = await db.inventory_activities.find_one(
        {"id": payload.activity_id}, {"_id": 0, "id": 1, "nama_kegiatan": 1})
    if not act:
        raise HTTPException(status_code=404, detail="Kegiatan inventarisasi tidak ditemukan")
    # Draft dibuat ke kegiatan ini — pastikan milik satker admin.
    await pastikan_akses_kegiatan_id(user, payload.activity_id)

    kategori_by_kode = {
        c["kode_aset"]: c.get("label", "")
        async for c in db.categories.find({}, {"_id": 0, "kode_aset": 1, "label": 1})
        if c.get("kode_aset")}

    maks = min(max(1, payload.maks), 1000)
    dibuat, dilewati_sudah_ada, gagal = 0, 0, []
    now = datetime.now(timezone.utc).isoformat()

    # Pasangan (kode barang, NUP) yang SUDAH tercatat — diambil SEKALI di luar
    # perulangan. Sebelumnya tiap baris memanggil `find_one` sendiri: 1.000
    # baris = 1.000 perjalanan bolak-balik ke MongoDB hanya untuk memutuskan
    # "lewati atau tidak", padahal satu kueri `$in` menjawab semuanya. Pola ini
    # menyalin cara `imports.py` menutup N+1 yang sama.
    pasangan = {(b.get("kode_barang", ""), b.get("nup", "")) for b in baris}
    pasangan.discard(("", ""))
    kode_semua = sorted({k for k, _ in pasangan if k})
    sudah_tercatat = set()
    if kode_semua:
        # Saring dulu dengan `$in` pada kode barang (ber-indeks), baru cocokkan
        # pasangan lengkapnya di Python — jauh lebih murah daripada `$or`
        # sepanjang ribuan klausa. Yang ditukar: himpunan ini memuat SEMUA aset
        # ber-kode sama, bukan hanya yang diminta. Proyeksinya dua field, jadi
        # 50.000 aset ≈ 2,5 MB — masih jauh lebih murah daripada 1.000
        # perjalanan bolak-balik, dan jumlah kode barang unik (bukan jumlah
        # baris) yang menentukan panjang `$in`.
        #
        # NUP dinormalkan ke teks di kedua sisi. Model menyimpannya sebagai
        # `Optional[str]`, tetapi data lama hasil impor bisa berupa angka —
        # `find_one` versi sebelumnya membandingkannya mentah, sehingga NUP 5
        # (angka) di basis data tak cocok dengan "5" dari berkas dan barisnya
        # dibuat DUA KALI.
        async for a in db.assets.find(
            {"asset_code": {"$in": kode_semua}, "dihapus": {"$ne": True}},
            {"_id": 0, "asset_code": 1, "NUP": 1},
        ):
            sudah_tercatat.add((str(a.get("asset_code") or ""), str(a.get("NUP") or "")))

    for b in baris:
        if dibuat >= maks:
            break
        kode, nup = b.get("kode_barang", ""), b.get("nup", "")
        if not kode or not nup:
            continue
        # Lewati bila kini sudah tercatat (impor ulang / dibuat manual).
        if (str(kode), str(nup)) in sudah_tercatat:
            dilewati_sudah_ada += 1
            continue
        # Baris kembar DI DALAM berkas yang sama juga harus dilewati: tanpa ini
        # daftar di atas tak ikut bertambah saat draft dibuat, sehingga dua
        # baris ber-NUP sama melahirkan dua aset.
        sudah_tercatat.add((str(kode), str(nup)))
        draft = AssetCreate(
            asset_code=kode,
            NUP=nup,
            asset_name=b.get("nama_barang") or f"Aset SIMAN {kode}",
            # Kategori = label kodefikasi lokal; fallback Nama Barang SIMAN
            # (uraian kodefikasi resmi) agar impor ulang tidak menandai selisih.
            category=kategori_by_kode.get(kode) or b.get("nama_barang", ""),
            brand=b.get("merk", ""),
            model=b.get("tipe", ""),
            condition=b.get("kondisi") or "Baik",
            purchase_price=str(int(round(float(b.get("nilai_perolehan") or 0)))),
            purchase_date=b.get("tanggal_perolehan", ""),
            kode_register=b.get("kode_register", ""),
            notes=f"Draft otomatis dari impor SIMAN V2 ({reg.get('filename', '')})",
            activity_id=payload.activity_id,
        )
        try:
            doc = await buat_aset_draft(
                draft, audit_user=user.get("name") or user.get("username") or "system")
            # Tandai langsung tersinkron SIMAN (sumber datanya memang SIMAN).
            await db.assets.update_one(
                {"id": doc["id"]},
                {"$set": {"siman": {"status": "cocok", "selisih": [],
                                    "kode_register": b.get("kode_register", ""),
                                    "import_id": import_id,
                                    "diperiksa_pada": now}}})
            # Jurnal Buku Barang (temuan selisih mutasi vs SIMAN): aset yang
            # lahir lewat jalur SIMAN dulu TANPA jurnal sama sekali — Buku
            # Barang kosong sampai backfill manual. Kini langsung 100 (Saldo
            # Awal) bertanggal buku = tanggal perolehan SIMAN. Best-effort +
            # anti-ganda via ref_id (pola create_asset).
            from shared_utils import catat_mutasi_bmn
            await catat_mutasi_bmn({
                "asset_id": doc["id"], "kode_transaksi": "100",
                "kode_barang": kode, "nup": nup,
                "tanggal_buku": (str(b.get("tanggal_perolehan") or "")[:10]
                                 or now[:10]),
                "jumlah": 1, "nilai": float(b.get("nilai_perolehan") or 0),
                "sumber_modul": "siman", "ref_id": doc["id"],
                "keterangan": ("Saldo awal dari impor SIMAN V2 "
                               f"({reg.get('filename', '')})"),
                "oleh": user.get("username", "system")})
            dibuat += 1
        except HTTPException as e:
            gagal.append(f"{kode}·{nup}: {e.detail}")
            if len(gagal) >= 20:
                break

    await log_audit("siman_buat_draft", payload.activity_id,
                    username=user.get("username", "system"),
                    detail=(f"Buat draft dari impor SIMAN '{reg.get('filename', '')}': "
                            f"{dibuat} dibuat, {dilewati_sudah_ada} sudah ada, "
                            f"{len(gagal)} gagal"))
    return {"dibuat": dibuat, "dilewati_sudah_ada": dilewati_sudah_ada,
            "gagal": gagal, "sisa": max(0, len(baris) - dibuat - dilewati_sudah_ada),
            "kegiatan": act.get("nama_kegiatan", "")}


@siman_router.get("/siman/selisih")
async def daftar_selisih_siman(page: int = 1, page_size: int = 50,
                               activity_id: str = "",
                               _user: dict = Depends(require_user)):
    """Aset yang datanya berbeda dengan SIMAN (untuk tabel tinjau & terapkan).

    `activity_id` menyaring ke SATU kegiatan inventarisasi — pasangan dari rekap
    per kegiatan di ringkasan, supaya "8 selisih di Inventarisasi Semester I"
    bisa langsung dibuka sebagai daftar, bukan dicari manual di antara seluruh
    selisih satker. Nilai "-" berarti aset yang belum terikat kegiatan mana pun.

    Penyaringan ditambahkan SETELAH `scope_query_aset`, jadi ia mempersempit
    lingkup — tak pernah bisa dipakai menembus batas satker.
    """
    page = max(1, page)
    page_size = min(max(1, page_size), 200)
    q = await scope_query_aset(_user, active_asset_filter({"siman.status": "selisih"}))
    pilih = (activity_id or "").strip()
    # SELALU lewat `$and`, JANGAN pernah menulis kunci "activity_id" langsung.
    # `scope_query_aset` menegakkan isolasi satker justru DENGAN kunci itu
    # (`activity_id: {"$in": [kegiatan milik satker user]}`), sehingga
    # `{**q, "activity_id": pilih}` menimpanya dan mengubah penyaring tampilan
    # menjadi IDOR: pengguna satker A cukup menyebut id kegiatan satker B untuk
    # membaca asetnya. Versi pertama kode ini melakukannya persis, dan uji
    # `test_penyaring_tak_bisa_menembus_batas_satker` yang menangkapnya.
    # Lewat `$and`, kedua syarat berlaku bersama → penyaring hanya bisa
    # MEMPERSEMPIT lingkup yang sudah diizinkan.
    if pilih == "-":
        q = {**q, "$and": [*q.get("$and", []),
                           {"$or": [{"activity_id": {"$in": ["", None]}},
                                    {"activity_id": {"$exists": False}}]}]}
    elif pilih:
        q = {**q, "$and": [*q.get("$and", []), {"activity_id": pilih}]}
    total = await db.assets.count_documents(q)
    items = await (db.assets.find(q, _PROJ_ASET)
                   .sort([("asset_code", 1), ("NUP", 1)])
                   .skip((page - 1) * page_size).limit(page_size)
                   .to_list(page_size))
    # Nama kegiatan disematkan di sini supaya tiap baris bisa berdiri sendiri:
    # tanpa ini daftar gabungan tak menunjukkan aset ini milik kegiatan mana.
    ids = {i.get("activity_id") for i in items if i.get("activity_id")}
    if ids:
        nama = {a["id"]: a.get("nama_kegiatan") or "" async for a in
                db.inventory_activities.find({"id": {"$in": list(ids)}},
                                             {"_id": 0, "id": 1, "nama_kegiatan": 1})}
        for i in items:
            i["nama_kegiatan"] = nama.get(i.get("activity_id"), "")
    return {"items": items, "total": total, "page": page,
            "activity_id": pilih,
            "total_pages": max(1, -(-total // page_size))}


@siman_router.post("/siman/terapkan/{asset_id}")
async def terapkan_siman(asset_id: str, payload: TerapkanIn,
                         user: dict = Depends(require_writer)):
    """Terapkan nilai SIMAN ke aset AMAN (menyingkronkan kembali).

    Default seluruh selisih; `fields` membatasi field tertentu. Nilai SIMAN
    menang (data valid). Ber-audit + $inc version (OCC).
    """
    a = await db.assets.find_one({"id": asset_id, "dihapus": {"$ne": True}},
                                 {"_id": 0, "id": 1, "NUP": 1, "asset_code": 1,
                                  "asset_name": 1, "activity_id": 1, "siman": 1,
                                  "version": 1})
    if not a:
        raise HTTPException(status_code=404, detail="Aset tidak ditemukan")
    await pastikan_akses_aset(user, a)
    sub = a.get("siman") or {}
    selisih = sub.get("selisih") or []
    if not selisih:
        raise HTTPException(status_code=400, detail="Tidak ada selisih SIMAN pada aset ini")

    pilih = set(payload.fields) if payload.fields else None
    terapkan = [s for s in selisih
                if s.get("field") in FIELD_TERAPKAN
                and (pilih is None or s.get("field") in pilih)]
    if not terapkan:
        raise HTTPException(status_code=400, detail="Tidak ada field terpilih yang bisa diterapkan")

    set_fields, changes = {}, []
    for s in terapkan:
        set_fields[s["field"]] = nilai_terapkan(s)
        changes.append({"field": s["field"], "from": s.get("aman", ""),
                        "to": nilai_terapkan(s)})

    # Perubahan kode barang = REKLASIFIKASI — tidak boleh lewat jalur
    # timpa-sinkron (menghapus jejak): wajib via POST /pembukuan/reklasifikasi
    # agar jurnal 304/107, riwayat_reklasifikasi, dan NUP tujuan tercatat
    # (integrasi audit #5; kartu sinkron menyediakan tombol khususnya).
    if "asset_code" in set_fields:
        raise HTTPException(status_code=409, detail=(
            "Perubahan kode barang adalah REKLASIFIKASI — gunakan tombol "
            "Reklasifikasi pada kartu sinkron SIMAN (jurnal 304/107 + riwayat "
            "tercatat). Field lain tetap bisa diterapkan dengan memilih field "
            "selain kode barang."))

    sisa = [s for s in selisih if s not in terapkan]
    sub.update({"selisih": sisa, "status": "selisih" if sisa else "cocok",
                "disinkron_pada": datetime.now(timezone.utc).isoformat()})
    now = datetime.now(timezone.utc).isoformat()
    res = await db.assets.find_one_and_update(
        {"id": asset_id, "dihapus": {"$ne": True}},
        {"$set": {**set_fields, "siman": sub, "updated_at": now},
         "$inc": {"version": 1}},
        projection={"_id": 0, "version": 1})
    if res is None:
        raise HTTPException(status_code=404, detail="Aset tidak ditemukan")

    # Jurnal 204/205 (temuan selisih mutasi vs SIMAN): menimpa purchase_price
    # dengan nilai SIMAN adalah KOREKSI NILAI — dulu senyap tanpa jejak Buku
    # Barang, padahal justru transaksi inilah padanan kolom "Nilai Mutasi"
    # SIMAN. Best-effort + anti-ganda via ref_id (pola penilaian.py).
    if "purchase_price" in set_fields:
        from pembukuan_utils import parse_harga
        from shared_utils import catat_mutasi_bmn
        lama = next((parse_harga(c["from"]) for c in changes
                     if c["field"] == "purchase_price"), 0.0)
        baru = parse_harga(set_fields["purchase_price"])
        selisih_nilai = baru - lama
        if abs(selisih_nilai) > 0.5:
            await catat_mutasi_bmn({
                "asset_id": asset_id,
                "kode_transaksi": "204" if selisih_nilai > 0 else "205",
                "kode_barang": str(a.get("asset_code") or ""),
                "nup": str(a.get("NUP") or ""),
                "tanggal_buku": now[:10],
                "jumlah": 0, "nilai": abs(selisih_nilai),
                "sumber_modul": "siman",
                "ref_id": f"siman-terapkan:{asset_id}:{int(lama)}:{int(baru)}",
                "keterangan": (f"Terapkan nilai SIMAN V2 — nilai perolehan "
                               f"Rp{int(lama):,} menjadi Rp{int(baru):,}"),
                "oleh": user.get("username", "system")})

    await log_audit("sinkron_siman", a.get("activity_id", ""), asset_id,
                    asset_code=a.get("asset_code", ""),
                    asset_name=a.get("asset_name", ""),
                    nup=str(a.get("NUP") or ""),
                    username=user.get("username", "system"),
                    changes=changes,
                    detail=f"Terapkan nilai SIMAN V2 ({len(terapkan)} field)")
    return {"ok": True, "diterapkan": [s["field"] for s in terapkan],
            "sisa_selisih": len(sisa), "version": (res or {}).get("version")}
