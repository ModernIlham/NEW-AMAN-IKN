"""Referensi kodefikasi barang BMN — fondasi Fase 2 (dipakai semua modul).

Koleksi `kodefikasi`: {kode, uraian, level (1-5), parent_kode} — struktur
kode diturunkan dari panjang prefix (lihat kodefikasi_utils.py). Dipakai
modul Pembukuan (DBKP per golongan), Persediaan (kode berawalan '1'),
Pengadaan (auto-golongan saat pendaftaran aset), dan laporan.

Endpoint baca terbuka untuk semua user login; tulis (CRUD/impor) admin.
Seed 8 golongan standar dilakukan idempoten saat koleksi masih kosong.
"""
import csv as csv_module
import io

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel, Field

from auth_utils import require_admin, require_user
from db import db
from kodefikasi_utils import (
    GOLONGAN_DEFAULTS, LEVEL_LABELS, META_FIELDS, derive_level,
    hierarchy_prefixes, is_persediaan_kode, normalize_kode, parent_of,
    parse_import_rows, validate_kode,
)

kodefikasi_router = APIRouter()


class KodefikasiIn(BaseModel):
    kode: str = Field(min_length=1, max_length=10)
    uraian: str = Field(min_length=1, max_length=300)


class KodefikasiUpdate(BaseModel):
    uraian: str = Field(min_length=1, max_length=300)


async def _ensure_golongan_seed():
    """Isi 8 golongan standar bila koleksi kosong — idempoten & murah."""
    if await db.kodefikasi.estimated_document_count() > 0:
        return
    for kode, uraian in GOLONGAN_DEFAULTS:
        await db.kodefikasi.update_one(
            {"kode": kode},
            {"$setOnInsert": {
                "kode": kode, "uraian": uraian, "level": 1, "parent_kode": None,
            }},
            upsert=True,
        )


def _doc(item: dict) -> dict:
    out = {
        "kode": item["kode"],
        "uraian": item.get("uraian", ""),
        "level": item.get("level"),
        "label_level": LEVEL_LABELS.get(item.get("level"), ""),
        "parent_kode": item.get("parent_kode"),
        "is_persediaan": is_persediaan_kode(item["kode"]),
    }
    # Metadata SIMAN (Satuan/Dasar/Jenis BMN/TB-STB/Bukti Kepemilikan) —
    # disajikan untuk panel Detail, bukan tabel utama.
    out["meta"] = {f: item.get(f, "") for f in META_FIELDS}
    return out


@kodefikasi_router.get("/kodefikasi")
async def list_kodefikasi(
    search: str = "",
    level: int = Query(0, ge=0, le=5),
    parent: str = "",
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    _user: dict = Depends(require_user),
):
    """Daftar kodefikasi: cari prefix kode / uraian, filter level/induk."""
    await _ensure_golongan_seed()
    query = {}
    if search:
        import re as _re
        s = search.strip()
        s_esc = _re.escape(s)
        query["$or"] = [
            {"kode": {"$regex": f"^{s_esc}" if s.isdigit() else s_esc, "$options": "i"}},
            {"uraian": {"$regex": s_esc, "$options": "i"}},
        ]
    if level:
        query["level"] = level
    if parent:
        query["parent_kode"] = normalize_kode(parent)
    total = await db.kodefikasi.count_documents(query)
    cursor = (db.kodefikasi.find(query, {"_id": 0})
              .sort("kode", 1).skip((page - 1) * page_size).limit(page_size))
    items = [_doc(x) async for x in cursor]
    return {
        "items": items, "total": total, "page": page, "page_size": page_size,
        "total_pages": max(1, -(-total // page_size)),
    }


@kodefikasi_router.get("/kodefikasi/template")
async def download_template(_user: dict = Depends(require_user)):
    """Template CSV impor: kolom kode,uraian — level diturunkan otomatis."""
    buf = io.StringIO()
    w = csv_module.writer(buf)
    w.writerow(["kode", "uraian"])
    w.writerow(["3", "Peralatan dan Mesin"])
    w.writerow(["301", "Alat Besar"])
    w.writerow(["30102", "Alat Besar Apung"])
    w.writerow(["3010203", "Contoh Sub Kelompok"])
    w.writerow(["3010203001", "Contoh Sub-sub Kelompok"])
    return Response(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="template_kodefikasi.csv"'},
    )


@kodefikasi_router.get("/kodefikasi/export")
async def export_kodefikasi(bentuk: str = Query("datar", pattern="^(datar|hierarki)$"),
                            _user: dict = Depends(require_user)):
    """Ekspor referensi kodefikasi (XLSX) — dua pendekatan:

    - **datar**: satu baris per kode (Kode, Uraian, Level, Kode Induk) — mengikuti
      tampilan tabel sekarang.
    - **hierarki**: satu baris per **kode barang level 5**, dengan kolom hierarki
      tiap tingkat (Golongan → Sub-Sub Kelompok) + metadata SIMAN (Satuan, Dasar,
      Jenis BMN, TB/STB, Bukti Kepemilikan).
    """
    from openpyxl import Workbook

    await _ensure_golongan_seed()
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Kodefikasi")

    if bentuk == "datar":
        ws.append(["Kode", "Uraian", "Level", "Kode Induk"])
        cursor = db.kodefikasi.find({}, {"_id": 0}).sort("kode", 1)
        async for x in cursor:
            lv = x.get("level")
            ws.append([x.get("kode", ""), x.get("uraian", ""),
                       f"{lv} {LEVEL_LABELS.get(lv, '')}".strip(),
                       x.get("parent_kode") or ""])
        nama_file = "kodefikasi_datar.xlsx"
    else:  # hierarki
        # Muat semua kode → uraian sekali untuk mengisi kolom leluhur.
        uraian_map = {}
        async for x in db.kodefikasi.find({}, {"_id": 0, "kode": 1, "uraian": 1}):
            uraian_map[x["kode"]] = x.get("uraian", "")
        ws.append([
            "Kode Golongan", "Nama Golongan", "Kode Bidang", "Nama Bidang",
            "Kode Kelompok", "Nama Kelompok", "Kode Sub Kelompok", "Nama Sub Kelompok",
            "Kode Barang", "Nama Barang", "Satuan", "Dasar", "Jenis BMN",
            "TB/STB", "Bukti Kepemilikan",
        ])
        cursor = db.kodefikasi.find({"level": 5}, {"_id": 0}).sort("kode", 1)
        async for x in cursor:
            k = x.get("kode", "")
            g, b, kl, sk = k[:1], k[:3], k[:5], k[:7]
            ws.append([
                g, uraian_map.get(g, ""), b, uraian_map.get(b, ""),
                kl, uraian_map.get(kl, ""), sk, uraian_map.get(sk, ""),
                k, x.get("uraian", ""),
                x.get("satuan", ""), x.get("dasar", ""), x.get("jenis_bmn", ""),
                x.get("tb_stb", ""), x.get("bukti_kepemilikan", ""),
            ])
        nama_file = "kodefikasi_hierarki.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nama_file}"'},
    )


@kodefikasi_router.get("/kodefikasi/lookup/{kode}")
async def lookup_kode(kode: str, _user: dict = Depends(require_user)):
    """Uraian berjenjang untuk sebuah kode barang (golongan → … → level kode).

    Menerima kode 1-10 digit ATAU lebih panjang (mis. kode persediaan 16
    digit — hanya 10 digit pertama yang berjenjang). Level yang uraiannya
    belum terdaftar tetap dikembalikan dengan uraian kosong.
    """
    kode = normalize_kode(kode)
    if not kode.isdigit():
        raise HTTPException(status_code=400, detail="Kode harus angka")
    await _ensure_golongan_seed()
    prefixes = hierarchy_prefixes(kode[:10])
    if not prefixes:
        raise HTTPException(status_code=400, detail="Kode terlalu pendek")
    kode_list = [p for _, p in prefixes]
    found = {x["kode"]: x async for x in db.kodefikasi.find(
        {"kode": {"$in": kode_list}}, {"_id": 0})}
    jenjang = [{
        "level": level,
        "label": LEVEL_LABELS[level],
        "kode": prefix,
        "uraian": found.get(prefix, {}).get("uraian", ""),
    } for level, prefix in prefixes]
    return {
        "kode": kode,
        "is_persediaan": is_persediaan_kode(kode),
        "jenjang": jenjang,
        "uraian_terdalam": next(
            (j["uraian"] for j in reversed(jenjang) if j["uraian"]), ""),
    }


@kodefikasi_router.post("/kodefikasi")
async def create_kodefikasi(data: KodefikasiIn, _admin: dict = Depends(require_admin)):
    kode = normalize_kode(data.kode)
    ok, err = validate_kode(kode)
    if not ok:
        raise HTTPException(status_code=400, detail=err)
    if await db.kodefikasi.find_one({"kode": kode}, {"_id": 1}):
        raise HTTPException(status_code=409, detail=f"Kode {kode} sudah terdaftar")
    doc = {
        "kode": kode, "uraian": data.uraian.strip(),
        "level": derive_level(kode), "parent_kode": parent_of(kode),
    }
    await db.kodefikasi.insert_one({**doc})
    return _doc(doc)


@kodefikasi_router.put("/kodefikasi/{kode}")
async def update_kodefikasi(kode: str, data: KodefikasiUpdate, _admin: dict = Depends(require_admin)):
    kode = normalize_kode(kode)
    res = await db.kodefikasi.update_one(
        {"kode": kode}, {"$set": {"uraian": data.uraian.strip()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail=f"Kode {kode} tidak ditemukan")
    item = await db.kodefikasi.find_one({"kode": kode}, {"_id": 0})
    return _doc(item)


@kodefikasi_router.delete("/kodefikasi/{kode}")
async def delete_kodefikasi(kode: str, _admin: dict = Depends(require_admin)):
    kode = normalize_kode(kode)
    child = await db.kodefikasi.find_one({"parent_kode": kode}, {"_id": 1})
    if child:
        raise HTTPException(
            status_code=409,
            detail=f"Kode {kode} masih punya turunan — hapus turunannya dulu")
    res = await db.kodefikasi.delete_one({"kode": kode})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail=f"Kode {kode} tidak ditemukan")
    return {"message": f"Kode {kode} dihapus"}


def _rows_from_upload(filename: str, content: bytes):
    """Baris dict dari CSV/XLSX — pola parser yang sama dengan imports.py."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        return list(csv_module.DictReader(io.StringIO(text)))
    if name.endswith(".xlsx") or name.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip().lower() for h in rows[0]]
        return [dict(zip(headers, r)) for r in rows[1:]]
    raise HTTPException(status_code=400, detail="Format file harus .csv atau .xlsx")


@kodefikasi_router.post("/kodefikasi/import")
async def import_kodefikasi(file: UploadFile = File(...), _admin: dict = Depends(require_admin)):
    """Impor massal CSV/XLSX aplikasi ATAU keluaran SIMAN V2 per level.

    Kode penuh & uraian dikenali dari header SIMAN (`Kode Barang`,
    `Nama Sub Subkelompok`, dll.); metadata SIMAN (Satuan/Dasar/Jenis BMN/
    TB-STB/Bukti Kepemilikan) ikut tersimpan. Kode sudah ada → diperbarui.
    """
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File maksimal 10MB")
    rows = _rows_from_upload(file.filename, content)
    entries, errors, dupes = parse_import_rows(rows)
    inserted = updated = meta_count = 0
    for e in entries:
        set_fields = {"uraian": e["uraian"], "level": e["level"],
                      "parent_kode": e["parent_kode"]}
        meta = {f: e[f] for f in META_FIELDS if f in e}
        if meta:
            set_fields.update(meta)
            meta_count += 1
        res = await db.kodefikasi.update_one(
            {"kode": e["kode"]},
            {"$set": set_fields, "$setOnInsert": {"kode": e["kode"]}},
            upsert=True,
        )
        if res.upserted_id is not None:
            inserted += 1
        elif res.modified_count:
            updated += 1
    return {
        "message": f"Impor selesai: {inserted} baru, {updated} diperbarui"
                   + (f", {meta_count} berinfo SIMAN" if meta_count else ""),
        "inserted": inserted, "updated": updated,
        "dengan_metadata": meta_count,
        "duplikat_dalam_file": dupes,
        "errors": errors[:50],
        "error_count": len(errors),
    }
