"""Parser impor Excel membaca lewat mode read-only openpyxl (TL-1).

Mode biasa membangun objek sel penuh untuk SELURUH sheet (~10x memori nilai
mentahnya) padahal kedua parser di sini hanya butuh nilai — pada berkas
mendekati batas unggah 15MB itu selisih ratusan MB per permintaan.

Mode read-only membawa jebakannya sendiri: ia MEMPERCAYAI metadata dimensi
di dalam file, dan metadata itu bisa berbohong ("mengaku 1 baris" — terbukti
pada ekspor SIMAN asli di siman.py). Karena itu `reset_dimensions()` wajib
menyertainya, dan uji di bawah memalsukan dimensi sungguhan untuk memastikan
penangkalnya terpasang — bukan sekadar memindai sumber.
"""
import io
import re
import zipfile

from routes.imports import parse_excel_content
from routes.kodefikasi import _rows_from_upload


def _xlsx(baris):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for r in baris:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _palsukan_dimensi(data):
    """Tulis ulang <dimension> sheet jadi "A1:A1" — meniru file nyata yang
    metadata dimensinya salah (ekspor SIMAN, sebagian generator pihak ketiga).
    """
    src = zipfile.ZipFile(io.BytesIO(data))
    out = io.BytesIO()
    tersulih = 0
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in src.infolist():
            isi = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                isi, n = re.subn(rb'<dimension ref="[^"]*"/>',
                                 rb'<dimension ref="A1:A1"/>', isi)
                tersulih += n
            zf.writestr(item.filename, isi)
    assert tersulih == 1, "bentuk XML openpyxl berubah; perbarui pemalsu ini"
    return out.getvalue()


class TestPerilakuTetapBenar:
    def test_parse_excel_content_membaca_semua_baris(self):
        rows, mulai = parse_excel_content(_xlsx([
            ["asset_code", "nama", "merk"],
            ["A-1", "Kursi", "Chitose"],
            ["A-2", "Meja", "Olympic"],
        ]))
        assert mulai == 2
        assert [r["asset_code"] for r in rows] == ["A-1", "A-2"]
        assert rows[0]["merk"] == "Chitose"

    def test_rows_from_upload_membaca_semua_baris(self):
        rows = _rows_from_upload("kode.xlsx", _xlsx([
            ["kode", "uraian"],
            ["1", "Tanah"],
            ["1.01", "Tanah Persil"],
        ]))
        assert [str(r["kode"]) for r in rows] == ["1", "1.01"]

    def test_baris_ragged_tidak_meledak(self):
        """Di mode read-only + reset_dimensions, sel kosong di EKOR baris
        tidak dikembalikan sama sekali (baris jadi lebih pendek dari header)
        — sebelumnya mode biasa mengembalikannya sebagai None. Seluruh
        konsumen kedua parser membaca lewat `.get(..., default)`, jadi kunci
        yang hilang setara dengan nilai kosong; uji ini mengunci asumsi itu.
        """
        rows, _ = parse_excel_content(_xlsx([
            ["asset_code", "nama", "merk"],
            ["A-1", "Kursi", None],
        ]))
        assert rows[0]["asset_code"] == "A-1"
        assert rows[0].get("merk", "") == ""
        rows2 = _rows_from_upload("k.xlsx", _xlsx([
            ["kode", "uraian", "satuan"],
            ["1", "Tanah", None],
        ]))
        assert rows2[0]["kode"] == "1"
        assert rows2[0].get("satuan") in (None, "")


class TestReadOnlyTerpasang:
    def test_kedua_parser_memakai_read_only(self, monkeypatch):
        """Bukan pemindaian sumber: sadap load_workbook dan periksa kwargs
        panggilan nyata — refactor apa pun yang mencabut read_only tertangkap.
        """
        import openpyxl
        dipanggil = []
        asli = openpyxl.load_workbook

        def mata(*a, **kw):
            dipanggil.append(dict(kw))
            return asli(*a, **kw)

        monkeypatch.setattr(openpyxl, "load_workbook", mata)
        data = _xlsx([["asset_code"], ["A-1"]])
        parse_excel_content(data)
        _rows_from_upload("x.xlsx", _xlsx([["kode", "uraian"], ["1", "T"]]))
        assert len(dipanggil) == 2, dipanggil
        for kw in dipanggil:
            assert kw.get("read_only") is True, kw
            assert kw.get("data_only") is True, kw


class TestDimensiBohong:
    """Inti kenapa read_only tidak boleh dipasang telanjang.

    Tanpa reset_dimensions, file yang mengaku "A1:A1" dibaca SATU SEL —
    impor 'berhasil' dengan 0 baris dan tak ada galat apa pun. Kedua uji ini
    gagal bila reset_dimensions dicabut dari parser masing-masing.
    """

    def test_parse_excel_content_kebal(self):
        data = _palsukan_dimensi(_xlsx([
            ["asset_code", "nama"],
            ["X-1", "Kursi"],
            ["X-2", "Meja"],
        ]))
        rows, _ = parse_excel_content(data)
        assert [r["asset_code"] for r in rows] == ["X-1", "X-2"]

    def test_rows_from_upload_kebal(self):
        data = _palsukan_dimensi(_xlsx([
            ["kode", "uraian"],
            ["1", "Tanah"],
            ["2", "Peralatan dan Mesin"],
        ]))
        rows = _rows_from_upload("k.xlsx", data)
        assert [str(r["kode"]) for r in rows] == ["1", "2"]
