"""Kerja CPU sinkron tidak boleh berjalan di event loop — temuan C26 & C27.

FastAPI melempar handler `def` biasa ke threadpool, tetapi handler `async def`
dijalankan LANGSUNG di event loop. Artinya satu panggilan Pillow atau satu
`workbook.close()` di dalam `async def` membekukan SELURUH proses selama ia
berjalan — bukan hanya permintaan itu.

Dua titik yang ditutup:

**C26 · `batch.py`** — `generate_photo_thumbnail`, `create_thumbnail`, dan
`create_gallery_thumbnail` dipanggil telanjang di dalam loop foto, sementara
fungsi yang SAMA sudah dibungkus benar di `assets.py:1199` dst. Diukur pada
JPEG 1600×1200: masing-masing ±37 ms dan ±36 ms. Delapan foto ±0,3 detik di
mesin uji. Yang dibeli di sini kerapian dan jitter yang hilang, bukan
penyelamatan darurat — dan uji ini menyebut itu apa adanya.

**C27 · `exports.py`** — `workbook.close()` bukan sekadar menutup berkas:
xlsxwriter merakit dan men-deflate SELURUH arsip di sana. Poin yang paling
mudah salah dipahami: memindahkannya ke "job latar" TIDAK menolong, sebab job
itu `asyncio.create_task` di event loop yang sama. "Latar" hanya berarti
pemanggilnya tak menunggu — loopnya tetap berhenti.

Penjaga di bawah memakai **AST, bukan grep**. Bedanya nyata: pola benar di
`assets.py` menaruh `lambda` pada baris BERIKUTNYA setelah `to_thread(`, jadi
penjaga per-baris akan melaporkannya sebagai pelanggaran. Grep tak bisa melihat
sarang; AST bisa.
"""
import ast
import pathlib

import pytest

BACKEND = pathlib.Path(__file__).resolve().parents[2]
ROUTES = BACKEND / "routes"

# Fungsi CPU sinkron yang WAJIB lewat thread bila dipanggil dari `async def`.
# Nama kartu dari C25b (decode Pillow + QR + ReportLab ±104 ms per kartu);
# `_rakit_pdf_ber_ttd` dari TL-3 (perakitan pypdf + ReportLab dokumen
# ber-TTD). HELPER PEMBUNGKUSNYA ikut masuk daftar — mutasi "cabut
# to_thread dari pemanggil helper" pernah SELAMAT karena penjaga hanya
# mengenal fungsi-dalamnya, sementara panggilan telanjang di coroutine
# berbentuk nama helper, bukan nama fungsi dalam.
SINKRON = {"generate_photo_thumbnail", "create_thumbnail",
           "create_gallery_thumbnail",
           "create_ktp_card_elements", "_draw_card_page",
           "_gambar_kartu_ke_kanvas", "_tutup_kanvas",
           "_rakit_pdf_ber_ttd"}


def _panggilan_telanjang(sumber: str, nama: set = None):
    """Nama fungsi SINKRON yang dipanggil di dalam `async def` tanpa to_thread.

    Mengembalikan [(nama, baris)]. Panggilan di dalam `def` biasa diabaikan:
    fungsi sinkron memanggil fungsi sinkron itu wajar — yang menentukan adalah
    siapa yang memanggil fungsi sinkron itu, dan itu urusan penjaga lain.

    `nama` bawaannya SINKRON (fungsi Pillow); penjaga C28 memakainya ulang
    dengan himpunan parser Excel — satu pendeteksi untuk dua kelas aturan,
    supaya perbaikan pada pendeteksinya menular ke keduanya.
    """
    nama = SINKRON if nama is None else nama
    pohon = ast.parse(sumber)

    # Petakan setiap simpul ke induknya sekali saja.
    induk = {}
    for n in ast.walk(pohon):
        for anak in ast.iter_child_nodes(n):
            induk[anak] = n

    def _dibungkus_to_thread(simpul) -> bool:
        """True bila simpul berada DI DALAM argumen `asyncio.to_thread(...)`."""
        n = induk.get(simpul)
        while n is not None:
            if isinstance(n, ast.Call):
                f = n.func
                if isinstance(f, ast.Attribute) and f.attr == "to_thread":
                    return True
                if isinstance(f, ast.Name) and f.id == "to_thread":
                    return True
            n = induk.get(n)
        return False

    def _dalam_async(simpul) -> bool:
        n = induk.get(simpul)
        while n is not None:
            if isinstance(n, ast.AsyncFunctionDef):
                return True
            if isinstance(n, ast.FunctionDef):
                return False        # fungsi sinkron bersarang — bukan urusan di sini
            n = induk.get(n)
        return False

    temuan = []
    for n in ast.walk(pohon):
        if not isinstance(n, ast.Call):
            continue
        # Nama polos (`load_workbook(...)`) DAN atribut (`openpyxl
        # .load_workbook(...)`) — parser yang sama dipanggil dua gaya di
        # repo ini, dan pendeteksi yang hanya melihat satu gaya setengah buta.
        if isinstance(n.func, ast.Name):
            f = n.func.id
        elif isinstance(n.func, ast.Attribute):
            f = n.func.attr
        else:
            f = None
        if f in nama and _dalam_async(n) and not _dibungkus_to_thread(n):
            temuan.append((f, n.lineno))
    return temuan


class TestPenjagaAstItuSendiri:
    """Penjaga yang tak pernah diuji pada kasus buruk bukan penjaga."""

    def test_menangkap_panggilan_telanjang(self):
        assert _panggilan_telanjang(
            "async def f():\n    x = create_thumbnail(b)\n") == [("create_thumbnail", 2)]

    def test_membebaskan_yang_dibungkus_to_thread(self):
        assert _panggilan_telanjang(
            "async def f():\n"
            "    x = await asyncio.to_thread(create_thumbnail, b)\n") == []

    def test_membebaskan_lambda_MULTIBARIS_di_dalam_to_thread(self):
        # Justru bentuk yang dipakai assets.py:1199 — dan justru bentuk yang
        # membuat penjaga berbasis grep melaporkan positif palsu.
        assert _panggilan_telanjang(
            "async def f():\n"
            "    x = await asyncio.to_thread(\n"
            "        lambda pics: [generate_photo_thumbnail(p) for p in pics], photos)\n") == []

    def test_mengabaikan_fungsi_sinkron_biasa(self):
        assert _panggilan_telanjang("def f():\n    return create_thumbnail(b)\n") == []

    def test_fungsi_sinkron_BERSARANG_di_async_juga_diabaikan(self):
        # Helper sinkron yang didefinisikan di dalam coroutine memang dibuat
        # untuk diserahkan ke to_thread; isinya bukan pelanggaran.
        assert _panggilan_telanjang(
            "async def f():\n"
            "    def bangun():\n"
            "        return create_thumbnail(b)\n"
            "    return await asyncio.to_thread(bangun)\n") == []


class TestSeluruhRoute:
    def test_tidak_ada_pillow_telanjang_di_coroutine_mana_pun(self):
        """Aturan KELAS, bukan tambalan satu berkas.

        C26 hanya menyebut batch.py, tetapi yang membuatnya terjadi adalah
        tidak adanya aturan — fungsi yang sama sudah dibungkus benar di
        assets.py sejak lama. Berkas berikutnya yang menyalin pola batch.py
        akan membuat CI merah, bukan lolos diam-diam.
        """
        pelanggar = {}
        for p in sorted(ROUTES.glob("*.py")):
            t = _panggilan_telanjang(p.read_text(encoding="utf-8"))
            if t:
                pelanggar[p.name] = t
        assert pelanggar == {}, pelanggar

    def test_batch_py_benar_benar_memakai_to_thread(self):
        # Penjaga di atas juga hijau bila seluruh panggilannya DIHAPUS.
        # Ini menagih fungsinya masih ada dan lewat thread. Hanya trio
        # thumbnail yang milik batch.py — nama kartu (C25b) diuji di
        # test_kartu_massal.py.
        src = (ROUTES / "batch.py").read_text(encoding="utf-8")
        assert src.count("asyncio.to_thread") >= 3, src.count("asyncio.to_thread")
        for f in ("generate_photo_thumbnail", "create_thumbnail",
                  "create_gallery_thumbnail"):
            assert f in src, f


class TestTutupWorkbook:
    SRC = (ROUTES / "exports.py").read_text(encoding="utf-8")

    def test_close_hanya_di_helper_sinkron(self):
        """`workbook.close()` tak boleh muncul lagi di dalam coroutine."""
        pohon = ast.parse(self.SRC)
        induk = {}
        for n in ast.walk(pohon):
            for a in ast.iter_child_nodes(n):
                induk[a] = n

        buruk = []
        for n in ast.walk(pohon):
            if not (isinstance(n, ast.Call) and isinstance(n.func, ast.Attribute)
                    and n.func.attr == "close"
                    and isinstance(n.func.value, ast.Name)
                    and n.func.value.id == "workbook"):
                continue
            p = induk.get(n)
            while p is not None:
                if isinstance(p, ast.AsyncFunctionDef):
                    buruk.append(n.lineno)
                    break
                if isinstance(p, ast.FunctionDef):
                    break
                p = induk.get(p)
        assert buruk == [], buruk

    def test_helpernya_sinkron_dan_dipanggil_lewat_thread(self):
        assert "def _tutup_dan_ambil(" in self.SRC
        assert "async def _tutup_dan_ambil(" not in self.SRC
        assert "await asyncio.to_thread(_tutup_dan_ambil" in self.SRC

    def test_helper_menghasilkan_XLSX_yang_sah(self):
        """Perbaikan performa tak boleh mengubah keluarannya."""
        import io
        import xlsxwriter
        from routes.exports import _tutup_dan_ambil

        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {"in_memory": True})
        wb.add_worksheet("Uji").write(0, 0, "halo")
        data = _tutup_dan_ambil(wb, buf)
        assert data[:2] == b"PK", data[:8]      # XLSX = arsip ZIP
        assert len(data) > 1000

    def test_helper_menyetel_ulang_posisi_buffer(self):
        # Tanpa seek(0), getvalue() memang tetap utuh — tetapi menghapus
        # seek-nya menandakan urutannya sudah tak dipahami lagi. Kunci
        # kontraknya: helper mengembalikan SELURUH isi, bukan ekornya.
        import io
        import xlsxwriter
        from routes.exports import _tutup_dan_ambil

        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {"in_memory": True})
        wb.add_worksheet("Uji").write(0, 0, "x")
        assert len(_tutup_dan_ambil(wb, buf)) == len(buf.getvalue())


class TestCatatanUtang:
    """Sisa parsing openpyxl di event loop — utang yang MENYUSUT.

    Sejarah daftar ini bagian dari nilainya. Ia sudah dua kali salah:

      1. `pegawai.py` terlewat saat disalin dari laporan — uji ini yang
         menangkapnya.
      2. `persediaan.py` masuk daftar padahal TIDAK menyebut openpyxl sama
         sekali. Ia memanggil `_rows_from_upload` yang di-impor dari
         `kodefikasi` — parser openpyxl yang sama, lewat pintu belakang.

    Temuan kedua mengungkap lubang di PENDETEKSINYA, bukan di daftarnya:
    mencari string "openpyxl" hanya melihat pemakaian LANGSUNG, sehingga
    pemanggil tak langsung sama-sama memblokir event loop dan sama-sama tak
    terlihat. Pendeteksi di bawah kini juga mengenali pemanggilan parser
    sinkron yang dikenal — dan justru itulah yang membuat `persediaan.py`
    ikut diperbaiki di C28, bukan tertinggal diam-diam.

    C28 menutup tiga berkas; tiga sisanya (siman, categories, pegawai —
    parsing menyatu dengan validasi/job sehingga butuh ekstraksi helper)
    ditutup di gelombang berikutnya, dan daftar ini KOSONG. Ia tetap ada
    sebagai penegak nol-utang: berkas baru yang menyentuh parser tanpa
    to_thread akan menabraknya. Daftar boleh menyusut, tak boleh bertambah —
    jadi ia SENGAJA diperketat tiap kali menyusut.
    """

    UTANG = []

    # Parser Excel sinkron yang dipakai lintas modul. Tiga nama terakhir
    # lahir dari ekstraksi C28: hadir di berkas mana pun tanpa to_thread =
    # utang baru.
    PARSER = ("openpyxl", "_rows_from_upload", "parse_excel_content",
              "_parse_siman_xlsx", "_parse_baris_impor_kategori",
              "_baris_dari_unggahan_pegawai")

    def _tanpa_thread(self):
        return [p.name for p in sorted(ROUTES.glob("*.py"))
                if any(k in (s := p.read_text(encoding="utf-8")) for k in self.PARSER)
                and "to_thread" not in s]

    def test_daftar_utang_tidak_bertambah(self):
        sisa = self._tanpa_thread()
        assert set(sisa) <= set(self.UTANG), sorted(set(sisa) - set(self.UTANG))

    def test_daftarnya_juga_tidak_boleh_KENDUR(self):
        """Entri yang sudah lunas harus DIKELUARKAN, bukan ditinggal menempel.

        Tanpa uji ini, daftar utang cuma tumbuh longgar: berkas yang sudah
        diperbaiki tetap terdaftar, dan penjaganya diam-diam mengizinkan
        berkas itu kembali rusak.
        """
        assert set(self.UTANG) == set(self._tanpa_thread()), (
            "daftar UTANG tak lagi cocok dengan kenyataan; perketat")


class TestC28:
    """Parser Excel sinkron dipanggil lewat thread di SEMUA titik panggil."""

    PARSER_TETAP_SINKRON = ["_rows_from_upload", "parse_excel_content",
                            "_parse_siman_xlsx", "_parse_baris_impor_kategori",
                            "_baris_dari_unggahan_pegawai",
                            # openpyxl langsung juga: endpoint impor BARU yang
                            # tak memakai parser di atas tetap tertangkap.
                            "load_workbook"]

    # (berkas, nama helper) hasil ekstraksi C28 — parsing yang dulu menyatu
    # dengan validasi/penulisan job di badan endpoint.
    HELPER = [("kodefikasi.py", "_rows_from_upload"),
              ("imports.py", "parse_excel_content"),
              ("siman.py", "_parse_siman_xlsx"),
              ("categories.py", "_parse_baris_impor_kategori"),
              ("pegawai.py", "_baris_dari_unggahan_pegawai")]

    @pytest.mark.parametrize("berkas,helper", HELPER)
    def test_parsernya_TETAP_sinkron(self, berkas, helper):
        """Sengaja tidak dijadikan `async def`.

        Semuanya parser murni tanpa I/O async. Menjadikannya async memaksa
        setiap pemanggil ikut berubah dan menyembunyikan biayanya di balik
        `await` yang tampak murah — padahal kerjanya tetap di event loop
        kecuali ada yang benar-benar memindahkannya ke thread.
        """
        s = (ROUTES / berkas).read_text(encoding="utf-8")
        assert f"def {helper}(" in s
        assert f"async def {helper}(" not in s

    def test_SEMUA_titik_panggil_lewat_thread(self):
        """Inti C28, dan bagian yang paling mudah setengah jalan.

        `_rows_from_upload` punya DUA pemanggil di dua berkas berbeda.
        Membungkusnya di `kodefikasi.py` saja meninggalkan `persediaan.py`
        tetap memblokir event loop — dan berkas itu tak pernah menyebut
        "openpyxl", jadi tak akan muncul di pencarian mana pun yang lazim.

        Memakai penjaga AST `_panggilan_telanjang` — BUKAN pemindaian
        per-baris. Versi pertama uji ini memindai per baris, dan itu kambuh
        ke persis pola yang dikritik docstring modul ini sendiri: bentuk
        yang SEMANTIKNYA BENAR seperti

            await asyncio.to_thread(
                lambda: parse_excel_content(content))

        akan DITOLAK karena `to_thread` dan nama parsernya jatuh di baris
        berbeda. Orang berikutnya yang membungkus siman.py dengan lambda
        multibaris akan membuat CI merah tanpa sebab — penjaga yang
        menghukum perbaikan yang benar lebih buruk daripada tak ada penjaga.
        """
        pelanggar = {}
        for p in sorted(ROUTES.glob("*.py")):
            t = _panggilan_telanjang(p.read_text(encoding="utf-8"),
                                     nama=set(self.PARSER_TETAP_SINKRON))
            if t:
                pelanggar[p.name] = t
        assert pelanggar == {}, pelanggar

    def test_kedua_pemanggil_rows_from_upload_masih_ada(self):
        # Penjaga di atas juga hijau bila pemanggilnya DIHAPUS seluruhnya.
        pemanggil = [p.name for p in sorted(ROUTES.glob("*.py"))
                     if "to_thread(_rows_from_upload" in p.read_text(encoding="utf-8")]
        assert set(pemanggil) == {"kodefikasi.py", "persediaan.py"}, pemanggil

    def test_imports_py_lewat_thread(self):
        src = (ROUTES / "imports.py").read_text(encoding="utf-8")
        assert "await asyncio.to_thread(parse_excel_content, content)" in src

    @pytest.mark.parametrize("berkas,helper", [
        ("siman.py", "_parse_siman_xlsx"),
        ("categories.py", "_parse_baris_impor_kategori"),
        ("pegawai.py", "_baris_dari_unggahan_pegawai"),
    ])
    def test_titik_panggil_tiga_berkas_terakhir_ada(self, berkas, helper):
        # Pasangan penjaga AST di atas: AST menjamin TIDAK ADA panggilan
        # telanjang, uji ini menjamin panggilan ber-thread-nya MASIH ADA —
        # tanpanya, menghapus fitur impornya sekalian juga lolos.
        s = (ROUTES / berkas).read_text(encoding="utf-8")
        assert f"to_thread({helper}" in s, berkas


class TestRateLimitImpor:
    """Semua endpoint impor massal ber-rate-limit — mitigasi ikutan C28.

    Setelah parsing pindah ke thread, blokirnya berpindah dari event loop ke
    threadpool default (min(32, cpu+4) = 8 di VPS 4 vCPU) yang DIPAKAI
    BERSAMA render PDF & thumbnail. Tanpa plafon, unggahan 10MB berulang
    menukar "membekukan semua" menjadi "menghabiskan threadpool" — lebih
    baik, tapi tetap penolakan layanan. Tiga endpoint pertama sudah lama
    berplafon; tiga terakhir baru dipasang bersama offload-nya.
    """

    ENDPOINT = [("imports.py", '.post("/import")'),
                ("siman.py", '.post("/siman/import")'),
                ("categories.py", '.post("/categories/import-bulk")'),
                ("kodefikasi.py", '.post("/kodefikasi/import")'),
                ("persediaan.py", '.post("/persediaan/import")'),
                ("pegawai.py", '.post("/pegawai/impor")')]

    @pytest.mark.parametrize("berkas,rute", ENDPOINT)
    def test_endpoint_impor_berplafon(self, berkas, rute):
        s = (ROUTES / berkas).read_text(encoding="utf-8")
        i = s.index(rute)
        # Dekorator limiter harus menempel di antara dekorator rute dan
        # `async def` — bukan sekadar hadir di suatu tempat dalam berkas.
        blok = s[i:s.index("async def", i)]
        assert "@limiter.limit(" in blok, f"{berkas} {rute} tanpa rate limit"


class TestPerilakuParserEkstraksi:
    """Ekstraksi C28 tidak boleh mengubah KELUARAN parser — hanya tempatnya.

    Penjaga struktural di atas menjamin bentuknya (sinkron + ber-thread);
    uji di sini menjamin isinya: xlsx kecil dibangun di memori, diparse lewat
    helper hasil ekstraksi, dan barisnya harus keluar utuh. Kalau ekstraksi
    diulang dan hasilnya berubah, berarti yang dipindah bukan parser murni.
    """

    @staticmethod
    def _xlsx(baris):
        import io as _io

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in baris:
            ws.append(r)
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_parse_siman_xlsx(self):
        import sys
        sys.path.insert(0, str(BACKEND))
        from routes.siman import _parse_siman_xlsx
        from tests.unit.test_siman_utils import HEADER_SIMAN, _baris

        isi = self._xlsx([["Kop Laporan"], HEADER_SIMAN,
                          _baris(), _baris(NUP="2")])
        # Parser mengembalikan (peta_header, baris, sheet) — peta & nama sheet
        # ikut dikembalikan karena pemanggilnya memakai keduanya. Sebelumnya
        # keduanya hanya variabel lokal di sini, dan handler yang membacanya
        # berakhir NameError → SETIAP impor SIMAN 500 (lihat
        # test_siman_import_endpoint.py).
        peta, hasil, sheet = _parse_siman_xlsx(isi)
        assert [b["nup"] for b in hasil] == ["1", "2"]
        assert hasil[0]["kode_barang"] == "3030203001"
        assert peta and sheet, "peta header / nama sheet tak ikut dikembalikan"

    def test_parse_siman_header_tak_dikenal_400(self):
        from fastapi import HTTPException

        from routes.siman import _parse_siman_xlsx
        with pytest.raises(HTTPException) as e:
            _parse_siman_xlsx(self._xlsx([["bukan", "header"]]))
        assert e.value.status_code == 400

    def test_parse_kategori_xlsx_dan_csv(self):
        from routes.categories import _parse_baris_impor_kategori

        rows = _parse_baris_impor_kategori("k.xlsx", self._xlsx(
            [["Kode Aset", "Deskripsi Barang"], ["3.01", "Alat Besar"]]))
        assert rows == [{"kode_aset": "3.01", "deskripsi_barang": "Alat Besar"}]

        rows = _parse_baris_impor_kategori(
            "k.csv", "Kode Aset,Deskripsi Barang\n3.02,Alat Angkut\n".encode())
        assert rows == [{"kode aset": "3.02", "deskripsi barang": "Alat Angkut"}]

    def test_parse_pegawai_xlsx_dan_format_salah(self):
        from fastapi import HTTPException

        from routes.pegawai import _baris_dari_unggahan_pegawai

        rows = _baris_dari_unggahan_pegawai("p.xlsx", self._xlsx(
            [["NIP", "Nama"], ["197001011990031001", "Budi"]]))
        assert rows == [{"NIP": "197001011990031001", "Nama": "Budi"}]
        with pytest.raises(HTTPException) as e:
            _baris_dari_unggahan_pegawai("p.pdf", b"x")
        assert e.value.status_code == 400
