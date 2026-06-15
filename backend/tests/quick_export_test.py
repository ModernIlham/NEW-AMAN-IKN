"""
Quick CSV/XLSX Export Test Script for Iteration 78
Verifies new jumlah_foto and tanggal_input columns
"""
import requests
import io
import sys

BASE_URL = "https://asset-crud-auth.preview.emergentagent.com"
ACTIVITY_ID = "2dad75d1-c43f-4c5b-8aad-3c6b48cce584"

def test_csv_export():
    print("=" * 60)
    print("Testing CSV Export")
    print("=" * 60)
    
    try:
        response = requests.get(
            f"{BASE_URL}/api/export/csv?activity_id={ACTIVITY_ID}",
            stream=True,
            timeout=120
        )
        
        print(f"Status Code: {response.status_code}")
        print(f"Content-Type: {response.headers.get('Content-Type', 'N/A')}")
        
        if response.status_code != 200:
            print(f"FAILED: {response.text[:500]}")
            return False
        
        # Read first chunk to verify headers
        content = b""
        for chunk in response.iter_content(chunk_size=4096):
            content += chunk
            if len(content) > 8192:  # Only need first ~8KB to verify
                break
        
        text = content.decode('utf-8', errors='replace')
        lines = text.split('\n')
        
        # Find header line (skip comments)
        header_line = None
        data_line = None
        comment_lines = []
        
        for line in lines:
            if line.startswith('#'):
                comment_lines.append(line)
            elif line.strip() and header_line is None:
                header_line = line
            elif line.strip() and header_line is not None:
                data_line = line
                break
        
        print(f"\nFound {len(comment_lines)} comment lines")
        for c in comment_lines[:5]:  # Show first 5 comments
            print(f"  {c}")
        
        if header_line:
            headers = [h.strip('"').strip() for h in header_line.split(',')]
            print(f"\nTotal CSV columns: {len(headers)}")
            print(f"Last 5 headers: {headers[-5:]}")
            
            # Check for new columns
            if 'jumlah_foto' in headers:
                print(f"✓ 'jumlah_foto' found at index {headers.index('jumlah_foto')}")
            else:
                print("✗ 'jumlah_foto' NOT FOUND")
                return False
            
            if 'tanggal_input' in headers:
                print(f"✓ 'tanggal_input' found at index {headers.index('tanggal_input')}")
            else:
                print("✗ 'tanggal_input' NOT FOUND")
                return False
            
            # Verify 41 columns
            if len(headers) == 41:
                print("✓ CSV has expected 41 columns")
            else:
                print(f"⚠ CSV has {len(headers)} columns (expected 41)")
        else:
            print("✗ No header line found")
            return False
        
        return True
        
    except Exception as e:
        print(f"ERROR: {e}")
        return False

def test_xlsx_export():
    print("\n" + "=" * 60)
    print("Testing XLSX Export")
    print("=" * 60)
    
    try:
        print("Downloading XLSX (may take 30-60 seconds for 1926 assets)...")
        response = requests.get(
            f"{BASE_URL}/api/export/xlsx?activity_id={ACTIVITY_ID}",
            timeout=300  # 5 minutes timeout
        )
        
        print(f"Status Code: {response.status_code}")
        print(f"Content-Type: {response.headers.get('Content-Type', 'N/A')}")
        print(f"Content-Length: {len(response.content)} bytes")
        
        if response.status_code != 200:
            print(f"FAILED: {response.text[:500]}")
            return False
        
        # Load workbook
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        sheet_names = wb.sheetnames
        
        print(f"\nSheet Names: {sheet_names}")
        
        expected_sheets = ['Data Aset', 'Kelengkapan Dokumen', 'Data Kegiatan', 'Tim Inventarisasi']
        for sheet in expected_sheets:
            if sheet in sheet_names:
                print(f"✓ Sheet '{sheet}' found")
            else:
                print(f"✗ Sheet '{sheet}' NOT FOUND")
                return False
        
        # Check Data Aset headers
        print("\n--- Data Aset Sheet ---")
        ws = wb['Data Aset']
        headers = [cell.value for cell in ws[1] if cell.value]
        print(f"Total headers: {len(headers)}")
        print(f"Last 3 headers: {headers[-3:] if len(headers) >= 3 else headers}")
        
        if 'Jumlah Foto' in headers:
            print("✓ 'Jumlah Foto' found")
        else:
            print("✗ 'Jumlah Foto' NOT FOUND")
        
        if 'Tanggal Input' in headers:
            print("✓ 'Tanggal Input' found")
        else:
            print("✗ 'Tanggal Input' NOT FOUND")
        
        # Check Data Kegiatan sections
        print("\n--- Data Kegiatan Sheet ---")
        ws_kegiatan = wb['Data Kegiatan']
        all_values = []
        for row in ws_kegiatan.iter_rows(max_row=50):
            for cell in row:
                if cell.value:
                    all_values.append(str(cell.value))
        content = ' '.join(all_values)
        
        sections = ['INFORMASI KEGIATAN', 'SATUAN KERJA', 'PENANGGUNG JAWAB', 'BERITA ACARA']
        for section in sections:
            if section in content:
                print(f"✓ Section '{section}' found")
            else:
                print(f"✗ Section '{section}' NOT FOUND")
        
        # Check Tim Inventarisasi
        print("\n--- Tim Inventarisasi Sheet ---")
        ws_tim = wb['Tim Inventarisasi']
        all_tim_values = []
        for row in ws_tim.iter_rows(max_row=100):
            for cell in row:
                if cell.value:
                    all_tim_values.append(str(cell.value))
        tim_content = ' '.join(all_tim_values)
        
        teams = ['TIM INTI', 'TIM PEMBANTU', 'TIM PENELITI', 'TIM PENDUKUNG']
        for team in teams:
            if team in tim_content:
                print(f"✓ Team '{team}' found")
            else:
                print(f"✗ Team '{team}' NOT FOUND")
        
        # Check tabular headers
        table_headers = ['No', 'Nama', 'Jabatan', 'NIP', 'Peran']
        found = [h for h in table_headers if h in tim_content]
        print(f"Table headers found: {found}")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    csv_pass = test_csv_export()
    xlsx_pass = test_xlsx_export()
    
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"CSV Export:  {'PASS ✓' if csv_pass else 'FAIL ✗'}")
    print(f"XLSX Export: {'PASS ✓' if xlsx_pass else 'FAIL ✗'}")
    
    sys.exit(0 if (csv_pass and xlsx_pass) else 1)
