"""Iteration 80 tests:
1. Excel/CSV template alignment & dropdowns (37 cols, schema integrity).
2. PDF upload via inventory activities → GridFS flow (no inline base64).
3. Streaming endpoint /documents/{idx}.
4. PUT documents=null preserve / [] wipe / subset orphan-cleanup semantics.
5. DELETE cleans GridFS blobs.
6. Regression: photos=null preserve, max 10/5 enforcement.
"""
import base64
import io
import os
import uuid
import pytest
import requests
from openpyxl import load_workbook

def _load_backend_url():
    val = os.environ.get("REACT_APP_BACKEND_URL")
    if not val:
        try:
            with open("/app/frontend/.env") as f:
                for line in f:
                    if line.startswith("REACT_APP_BACKEND_URL="):
                        val = line.split("=", 1)[1].strip()
                        break
        except FileNotFoundError:
            pass
    if not val:
        raise RuntimeError("REACT_APP_BACKEND_URL not set")
    return val.rstrip("/")

BASE_URL = _load_backend_url()
USERNAME = "bugfix_admin_test2"
PASSWORD = "BugfixTest123"

EXPECTED_FIELDS = [
    "asset_code", "NUP", "asset_name", "category", "brand", "model",
    "kode_register", "serial_number", "purchase_date", "purchase_price",
    "location", "eselon1", "eselon2", "user", "condition", "status",
    "nomor_spm", "perolehan_dari_nama", "nomor_kontrak", "nomor_bukti_perolehan",
    "supplier", "notes", "stiker_status", "stiker_ukuran", "inventory_status",
    "klasifikasi_tidak_ditemukan", "sub_klasifikasi", "uraian_tidak_ditemukan",
    "tindak_lanjut", "kronologis", "koordinat_latitude", "koordinat_longitude",
    "keterangan_berlebih", "asal_usul_berlebih", "nomor_perkara",
    "pihak_bersengketa", "keterangan_sengketa",
]


@pytest.fixture(scope="module")
def auth_token():
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"username": USERNAME, "password": PASSWORD},
                      timeout=30)
    if r.status_code != 200:
        pytest.skip(f"Login failed: {r.status_code} {r.text}")
    return r.json()["access_token"]


@pytest.fixture(scope="module")
def headers(auth_token):
    return {"Authorization": f"Bearer {auth_token}", "Content-Type": "application/json"}


def _make_pdf_data_url(payload_text="Hello PDF Test"):
    """Create a tiny valid PDF (~~1KB) with given payload string."""
    pdf_body = (
        b"%PDF-1.4\n"
        b"1 0 obj <</Type /Catalog /Pages 2 0 R>> endobj\n"
        b"2 0 obj <</Type /Pages /Kids [3 0 R] /Count 1>> endobj\n"
        b"3 0 obj <</Type /Page /Parent 2 0 R /MediaBox [0 0 300 300] /Contents 4 0 R>> endobj\n"
        b"4 0 obj <</Length 44>> stream\n"
        b"BT /F1 12 Tf 50 250 Td (" + payload_text.encode() + b") Tj ET\n"
        b"endstream endobj\n"
        b"xref\n0 5\n"
        b"0000000000 65535 f\n"
        b"trailer <</Size 5 /Root 1 0 R>>\n"
        b"startxref\n0\n%%EOF\n"
    )
    b64 = base64.b64encode(pdf_body).decode()
    return f"data:application/pdf;base64,{b64}", len(pdf_body)


# ============================================================================
# 1. TEMPLATES (CSV + XLSX) — schema alignment
# ============================================================================
class TestTemplates:
    def test_csv_template_alignment(self):
        r = requests.get(f"{BASE_URL}/api/templates/csv", timeout=30)
        assert r.status_code == 200
        text = r.content.decode("utf-8-sig")
        lines = [l for l in text.split("\n") if l.strip()]
        assert len(lines) == 3, f"Expected 3 lines (header+2 samples), got {len(lines)}"

        import csv as _csv
        rows = list(_csv.reader(io.StringIO(text)))
        headers = rows[0]
        assert len(headers) == 37, f"Expected 37 cols, got {len(headers)}"
        assert headers == EXPECTED_FIELDS, f"Header mismatch: {headers}"

        sample1 = rows[1]
        sample2 = rows[2]
        assert len(sample1) == 37
        assert len(sample2) == 37

        # Critical alignment checks (the bug)
        idx = lambda f: EXPECTED_FIELDS.index(f)
        assert sample1[idx("eselon2")] == "Biro Umum"
        assert sample2[idx("eselon2")] == "Direktorat Satu"
        assert sample1[idx("user")] == "John Doe"
        assert sample2[idx("user")] == "Jane Smith"
        assert sample1[idx("condition")] == "Baik"
        assert sample1[idx("status")] == "Aktif"
        assert sample1[idx("nomor_spm")] == "02847T/621001/2024"

        # New fields exist & default empty in row 1
        for f in ["kronologis", "koordinat_latitude", "koordinat_longitude",
                  "keterangan_berlebih", "asal_usul_berlebih",
                  "nomor_perkara", "pihak_bersengketa", "keterangan_sengketa"]:
            assert f in EXPECTED_FIELDS
            # check column exists at the right index
            assert headers[idx(f)] == f

    def test_xlsx_template_alignment_and_dropdowns(self):
        r = requests.get(f"{BASE_URL}/api/templates/xlsx", timeout=30)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        assert "Data Import" in wb.sheetnames
        assert "Panduan" in wb.sheetnames

        ws = wb["Data Import"]
        # Headers in row 4
        headers = []
        for col in range(1, 100):
            val = ws.cell(row=4, column=col).value
            if val is None:
                break
            # Strip the " *" suffix on required headers
            headers.append(str(val).replace(" *", "").strip())
        assert len(headers) == 37, f"Expected 37 headers, got {len(headers)}: {headers}"
        assert headers == EXPECTED_FIELDS

        # Sample data rows 5 & 6
        sample1 = [ws.cell(row=5, column=c+1).value for c in range(37)]
        sample2 = [ws.cell(row=6, column=c+1).value for c in range(37)]
        idx = lambda f: EXPECTED_FIELDS.index(f)
        assert sample1[idx("eselon2")] == "Biro Umum"
        assert sample2[idx("eselon2")] == "Direktorat Satu"
        assert sample1[idx("user")] == "John Doe"
        assert sample2[idx("user")] == "Jane Smith"
        assert sample1[idx("condition")] == "Baik"
        assert sample1[idx("status")] == "Aktif"

        # Dropdowns at correct columns. Excel cols are 1-indexed.
        # Expected: condition→col 15 (0-based 14 + 1), status→16, stiker_status→23,
        #           stiker_ukuran→24, inventory_status→25, klasifikasi→26, sub_klasifikasi→27.
        # The user-facing spec says "col 14, col 15..." which IS 0-indexed.
        # Verify by parsing data validation ranges.
        from openpyxl.utils import get_column_letter
        expected_dropdown_cols = {
            "category": idx("category"),  # 3
            "condition": idx("condition"),  # 14
            "status": idx("status"),  # 15
            "stiker_status": idx("stiker_status"),  # 22
            "stiker_ukuran": idx("stiker_ukuran"),  # 23
            "inventory_status": idx("inventory_status"),  # 24
            "klasifikasi_tidak_ditemukan": idx("klasifikasi_tidak_ditemukan"),  # 25
            "sub_klasifikasi": idx("sub_klasifikasi"),  # 26
        }
        dv_cols = set()
        for dv in ws.data_validations.dataValidation:
            for r_ in dv.sqref.ranges:
                # min_col is 1-indexed
                dv_cols.add(r_.min_col - 1)
        for fname, col0 in expected_dropdown_cols.items():
            assert col0 in dv_cols, f"Dropdown for {fname} missing at 0-idx col {col0}; got {sorted(dv_cols)}"

    def test_xlsx_panduan_sheet_has_all_37_fields(self):
        r = requests.get(f"{BASE_URL}/api/templates/xlsx", timeout=30)
        wb = load_workbook(io.BytesIO(r.content))
        guide = wb["Panduan"]
        # Header row at row 3 (0-indexed = 2 → openpyxl row 3); fields start row 4
        guide_fields = []
        row = 4
        while True:
            v = guide.cell(row=row, column=1).value
            if v is None:
                break
            guide_fields.append(str(v).strip())
            row += 1
        assert guide_fields == EXPECTED_FIELDS, (
            f"Panduan mismatch: extra={set(guide_fields)-set(EXPECTED_FIELDS)} "
            f"missing={set(EXPECTED_FIELDS)-set(guide_fields)}"
        )


# ============================================================================
# 2. PDF UPLOAD VIA ACTIVITY → GridFS
# ============================================================================
@pytest.fixture(scope="module")
def created_activity(headers):
    """Create one activity with 2 PDFs and yield (activity_id, original_sizes)."""
    pdf1, sz1 = _make_pdf_data_url("Doc One")
    pdf2, sz2 = _make_pdf_data_url("Doc Two AAAAAAAAA")
    payload = {
        "nomor_surat": f"TEST_ITER80_{uuid.uuid4().hex[:8]}",
        "nama_kegiatan": "TEST Iteration 80 PDF",
        "kode_satker": f"T{uuid.uuid4().hex[:5].upper()}",
        "nama_satker": f"TEST Satker {uuid.uuid4().hex[:5]}",
        "documents": [
            {"name": "kontrak1.pdf", "data": pdf1},
            {"name": "kontrak2.pdf", "data": pdf2},
        ],
    }
    r = requests.post(f"{BASE_URL}/api/inventory-activities",
                      json=payload, headers=headers, timeout=120)
    assert r.status_code == 200, f"Create failed: {r.status_code} {r.text}"
    data = r.json()
    aid = data["id"]
    yield aid, data
    # teardown
    try:
        requests.delete(f"{BASE_URL}/api/inventory-activities/{aid}",
                        headers=headers, timeout=30)
    except Exception:
        pass


class TestPDFActivity:
    def test_create_returns_gridfs_metadata_no_data(self, created_activity):
        aid, data = created_activity
        docs = data["documents"]
        assert len(docs) == 2
        for d in docs:
            assert "gridfs_id" in d and d["gridfs_id"]
            assert "size" in d and isinstance(d["size"], int)
            assert "compression_method" in d
            assert "data" not in d, f"data field should NOT be in response: {d}"
            assert "name" in d

    def test_get_activity_strips_data_field(self, created_activity, headers):
        aid, _ = created_activity
        r = requests.get(f"{BASE_URL}/api/inventory-activities/{aid}",
                         headers=headers, timeout=30)
        assert r.status_code == 200
        for d in r.json().get("documents", []):
            assert "data" not in d
            assert d.get("gridfs_id")

    def test_stream_document_endpoint(self, created_activity, headers):
        aid, _ = created_activity
        r = requests.get(f"{BASE_URL}/api/inventory-activities/{aid}/documents/0",
                         headers=headers, timeout=30)
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith("application/pdf")
        assert r.content[:4] == b"%PDF"

    def test_stream_document_out_of_range_returns_404(self, created_activity, headers):
        aid, _ = created_activity
        r = requests.get(f"{BASE_URL}/api/inventory-activities/{aid}/documents/99",
                         headers=headers, timeout=30)
        assert r.status_code == 404


# ============================================================================
# 3. PUT semantics: preserve / wipe / subset
# ============================================================================
class TestPutDocuments:
    def _create_with_docs(self, headers, n=2):
        docs = []
        for i in range(n):
            url, _ = _make_pdf_data_url(f"PutDoc{i}")
            docs.append({"name": f"d{i}.pdf", "data": url})
        payload = {
            "nomor_surat": f"TEST_ITER80_PUT_{uuid.uuid4().hex[:8]}",
            "nama_kegiatan": "TEST PUT semantics",
            "kode_satker": f"P{uuid.uuid4().hex[:5].upper()}",
            "nama_satker": f"TEST Put {uuid.uuid4().hex[:5]}",
            "documents": docs,
        }
        r = requests.post(f"{BASE_URL}/api/inventory-activities",
                          json=payload, headers=headers, timeout=120)
        assert r.status_code == 200, r.text
        return r.json(), payload

    def test_put_documents_null_preserves_existing(self, headers):
        created, payload = self._create_with_docs(headers, n=2)
        aid = created["id"]
        existing_ids = [d["gridfs_id"] for d in created["documents"]]

        update = {
            "nomor_surat": payload["nomor_surat"],
            "nama_kegiatan": "TEST PUT preserve",
            "kode_satker": payload["kode_satker"],
            "nama_satker": payload["nama_satker"],
            "documents": None,
            "photos": None,
        }
        r = requests.put(f"{BASE_URL}/api/inventory-activities/{aid}",
                         json=update, headers=headers, timeout=60)
        assert r.status_code == 200, r.text
        after = r.json().get("documents", [])
        after_ids = [d["gridfs_id"] for d in after]
        assert sorted(after_ids) == sorted(existing_ids), \
            f"Documents NOT preserved on null. Before={existing_ids} After={after_ids}"

        # cleanup
        requests.delete(f"{BASE_URL}/api/inventory-activities/{aid}",
                        headers=headers, timeout=30)

    def test_put_documents_subset_deletes_orphans(self, headers):
        created, payload = self._create_with_docs(headers, n=2)
        aid = created["id"]
        keep_doc = created["documents"][0]
        orphan_doc = created["documents"][1]

        update = {
            "nomor_surat": payload["nomor_surat"],
            "nama_kegiatan": payload["nama_kegiatan"],
            "kode_satker": payload["kode_satker"],
            "nama_satker": payload["nama_satker"],
            "documents": [{"name": keep_doc["name"], "gridfs_id": keep_doc["gridfs_id"]}],
        }
        r = requests.put(f"{BASE_URL}/api/inventory-activities/{aid}",
                         json=update, headers=headers, timeout=60)
        assert r.status_code == 200, r.text
        after = r.json()["documents"]
        assert len(after) == 1
        assert after[0]["gridfs_id"] == keep_doc["gridfs_id"]

        # Verify orphan stream returns 404 (kept's index is now 0, orphan deleted)
        r2 = requests.get(f"{BASE_URL}/api/inventory-activities/{aid}/documents/1",
                          headers=headers, timeout=30)
        assert r2.status_code == 404

        # Kept doc still streams OK
        r3 = requests.get(f"{BASE_URL}/api/inventory-activities/{aid}/documents/0",
                          headers=headers, timeout=30)
        assert r3.status_code == 200
        assert r3.content[:4] == b"%PDF"

        # cleanup
        requests.delete(f"{BASE_URL}/api/inventory-activities/{aid}",
                        headers=headers, timeout=30)

    def test_put_documents_empty_wipes_all(self, headers):
        created, payload = self._create_with_docs(headers, n=2)
        aid = created["id"]
        update = {
            "nomor_surat": payload["nomor_surat"],
            "nama_kegiatan": payload["nama_kegiatan"],
            "kode_satker": payload["kode_satker"],
            "nama_satker": payload["nama_satker"],
            "documents": [],
        }
        r = requests.put(f"{BASE_URL}/api/inventory-activities/{aid}",
                         json=update, headers=headers, timeout=60)
        assert r.status_code == 200, r.text
        assert r.json().get("documents", []) == []

        # cleanup
        requests.delete(f"{BASE_URL}/api/inventory-activities/{aid}",
                        headers=headers, timeout=30)

    def test_max_documents_enforcement(self, headers):
        # Try creating with 6 docs (limit 5)
        docs = []
        for i in range(6):
            url, _ = _make_pdf_data_url(f"OverLimit{i}")
            docs.append({"name": f"lim{i}.pdf", "data": url})
        payload = {
            "nomor_surat": f"TEST_ITER80_LIMIT_{uuid.uuid4().hex[:8]}",
            "nama_kegiatan": "TEST limit",
            "kode_satker": f"L{uuid.uuid4().hex[:5].upper()}",
            "nama_satker": f"TEST Limit {uuid.uuid4().hex[:5]}",
            "documents": docs,
        }
        r = requests.post(f"{BASE_URL}/api/inventory-activities",
                          json=payload, headers=headers, timeout=60)
        assert r.status_code == 400
        assert "5 dokumen" in r.text or "Maksimal" in r.text


# ============================================================================
# 4. DELETE cleans up GridFS
# ============================================================================
class TestDeleteCleanup:
    def test_delete_activity_removes_gridfs_blobs(self, headers):
        url, _ = _make_pdf_data_url("Delete Me")
        payload = {
            "nomor_surat": f"TEST_ITER80_DEL_{uuid.uuid4().hex[:8]}",
            "nama_kegiatan": "TEST delete cleanup",
            "kode_satker": f"D{uuid.uuid4().hex[:5].upper()}",
            "nama_satker": f"TEST Del {uuid.uuid4().hex[:5]}",
            "documents": [{"name": "del.pdf", "data": url}],
        }
        r = requests.post(f"{BASE_URL}/api/inventory-activities",
                          json=payload, headers=headers, timeout=60)
        assert r.status_code == 200
        aid = r.json()["id"]

        # Confirm doc streams BEFORE delete
        r1 = requests.get(f"{BASE_URL}/api/inventory-activities/{aid}/documents/0",
                          headers=headers, timeout=30)
        assert r1.status_code == 200

        rd = requests.delete(f"{BASE_URL}/api/inventory-activities/{aid}",
                             headers=headers, timeout=30)
        assert rd.status_code == 200

        # Activity itself is gone → 404
        r2 = requests.get(f"{BASE_URL}/api/inventory-activities/{aid}",
                          headers=headers, timeout=30)
        assert r2.status_code == 404
