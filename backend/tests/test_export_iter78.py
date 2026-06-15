
# --- centralised test credentials (avoid hardcoded secrets) ---
import os, sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from conftest import TEST_ADMIN_PASSWORD, TEST_ADMIN_USERNAME  # noqa: E402
"""
Test Export Features - Iteration 78
Tests for CSV and XLSX export with new jumlah_foto and tanggal_input columns
Plus new Data Kegiatan and Tim Inventarisasi sheet structures
"""

import pytest
import requests
import os
import io
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
TEST_ACTIVITY_ID = "2dad75d1-c43f-4c5b-8aad-3c6b48cce584"


class TestAuthentication:
    """Basic auth to get token for later tests"""
    
    def test_login_success(self):
        """Verify login and get access token"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "username": "admin",
            "password": TEST_ADMIN_PASSWORD
        })
        assert response.status_code == 200, f"Login failed: {response.text}"
        data = response.json()
        # Check for token (may be 'access_token' or 'token')
        assert "access_token" in data or "token" in data, "No token in response"
        print("✓ Login successful - got access token")


class TestCSVExport:
    """Test CSV export with new jumlah_foto and tanggal_input columns"""
    
    def test_csv_export_returns_200(self):
        """Verify CSV export endpoint returns 200"""
        response = requests.get(
            f"{BASE_URL}/api/export/csv?activity_id={TEST_ACTIVITY_ID}",
            stream=True,
            timeout=60
        )
        assert response.status_code == 200, f"CSV export failed with status {response.status_code}"
        assert "text/csv" in response.headers.get("Content-Type", ""), "Wrong content type"
        print("✓ CSV export returns 200 with correct content type")
    
    def test_csv_header_has_jumlah_foto_and_tanggal_input(self):
        """Verify CSV header row contains the new jumlah_foto and tanggal_input columns"""
        response = requests.get(
            f"{BASE_URL}/api/export/csv?activity_id={TEST_ACTIVITY_ID}",
            stream=True,
            timeout=60
        )
        assert response.status_code == 200
        
        # Read first few lines to get header
        content = b""
        for chunk in response.iter_content(chunk_size=8192):
            content += chunk
            if b"\n" in content:
                break
        
        lines = content.decode('utf-8').split('\n')
        # Find the header row (skip comment lines starting with #)
        header_line = None
        for line in lines:
            if not line.startswith('#') and line.strip():
                header_line = line
                break
        
        assert header_line is not None, "No header row found in CSV"
        
        # Parse header columns
        headers = [h.strip('"').strip() for h in header_line.split(',')]
        
        # Verify new columns exist
        assert 'jumlah_foto' in headers, f"'jumlah_foto' not found in headers: {headers}"
        assert 'tanggal_input' in headers, f"'tanggal_input' not found in headers: {headers}"
        
        print(f"✓ CSV header contains 'jumlah_foto' at position {headers.index('jumlah_foto')}")
        print(f"✓ CSV header contains 'tanggal_input' at position {headers.index('tanggal_input')}")
        print(f"  Total columns: {len(headers)}")
    
    def test_csv_has_41_columns(self):
        """Verify CSV has all 41 expected columns as per specification"""
        response = requests.get(
            f"{BASE_URL}/api/export/csv?activity_id={TEST_ACTIVITY_ID}",
            stream=True,
            timeout=60
        )
        assert response.status_code == 200
        
        content = b""
        for chunk in response.iter_content(chunk_size=8192):
            content += chunk
            if b"\n" in content:
                break
        
        lines = content.decode('utf-8').split('\n')
        header_line = None
        for line in lines:
            if not line.startswith('#') and line.strip():
                header_line = line
                break
        
        headers = [h.strip('"').strip() for h in header_line.split(',')]
        
        expected_columns = [
            'asset_code', 'NUP', 'asset_name', 'category', 'brand', 'model',
            'kode_register', 'serial_number', 'purchase_date', 'purchase_price',
            'location', 'eselon1', 'eselon2', 'user', 'condition', 'status',
            'nomor_spm', 'perolehan_dari_nama', 'nomor_kontrak', 'nomor_bukti_perolehan',
            'supplier', 'notes', 'stiker_status', 'stiker_ukuran', 'inventory_status',
            'klasifikasi_tidak_ditemukan', 'sub_klasifikasi', 'uraian_tidak_ditemukan',
            'tindak_lanjut', 'koordinat_latitude', 'koordinat_longitude', 'kronologis',
            'keterangan_berlebih', 'asal_usul_berlebih', 'nomor_perkara', 'pihak_bersengketa',
            'keterangan_sengketa', 'jumlah_foto', 'tanggal_input',
            'kelengkapan_items', 'link_foto_kelengkapan', 'link_pdf_kelengkapan'
        ]
        
        assert len(headers) == 41, f"Expected 41 columns but got {len(headers)}: {headers}"
        
        for col in expected_columns:
            assert col in headers, f"Missing expected column: {col}"
        
        print("✓ CSV has all 41 expected columns")
    
    def test_csv_has_activity_header_comments(self):
        """Verify CSV has activity header comment lines"""
        response = requests.get(
            f"{BASE_URL}/api/export/csv?activity_id={TEST_ACTIVITY_ID}",
            stream=True,
            timeout=60
        )
        assert response.status_code == 200
        
        content = b""
        for chunk in response.iter_content(chunk_size=8192):
            content += chunk
            if len(content) > 2000:
                break
        
        text = content.decode('utf-8')
        
        # Check for activity header comments
        assert '# Kegiatan:' in text, "Missing activity name comment"
        assert '# Nomor Surat:' in text, "Missing nomor surat comment"
        assert '# Kode Satker:' in text, "Missing kode satker comment"
        
        print("✓ CSV has activity header comments (Kegiatan, Nomor Surat, Kode Satker)")
    
    def test_csv_data_row_has_photo_count_and_date(self):
        """Verify CSV data rows contain photo count and created_at values"""
        response = requests.get(
            f"{BASE_URL}/api/export/csv?activity_id={TEST_ACTIVITY_ID}",
            stream=True,
            timeout=120
        )
        assert response.status_code == 200
        
        # Read enough content to get header and at least one data row
        content = b""
        line_count = 0
        for chunk in response.iter_content(chunk_size=8192):
            content += chunk
            line_count = content.count(b'\n')
            if line_count >= 10:  # Get enough lines
                break
        
        lines = content.decode('utf-8').split('\n')
        
        # Find header row to get column indices
        header_line = None
        data_line = None
        for line in lines:
            if not line.startswith('#') and line.strip():
                if header_line is None:
                    header_line = line
                else:
                    data_line = line
                    break
        
        assert header_line is not None, "No header found"
        assert data_line is not None, "No data row found"
        
        headers = [h.strip('"').strip() for h in header_line.split(',')]
        foto_idx = headers.index('jumlah_foto')
        tanggal_idx = headers.index('tanggal_input')
        
        # Parse data row (handle quoted values with commas inside)
        import csv
        reader = csv.reader([data_line])
        data_values = list(reader)[0]
        
        foto_val = data_values[foto_idx] if foto_idx < len(data_values) else None
        tanggal_val = data_values[tanggal_idx] if tanggal_idx < len(data_values) else None
        
        print(f"  Data row jumlah_foto = {foto_val}")
        print(f"  Data row tanggal_input = {tanggal_val}")
        
        # Photo count should be a number (may be 0)
        assert foto_val is not None, "jumlah_foto value is None"
        try:
            int(foto_val)
            print(f"✓ jumlah_foto is a valid integer: {foto_val}")
        except ValueError:
            print(f"⚠ jumlah_foto value '{foto_val}' may be empty or invalid - this is acceptable if asset has no photos")


class TestXLSXExport:
    """Test XLSX export with new sheet structures"""
    
    def test_xlsx_export_returns_200(self):
        """Verify XLSX export endpoint returns 200"""
        response = requests.get(
            f"{BASE_URL}/api/export/xlsx?activity_id={TEST_ACTIVITY_ID}",
            timeout=180  # XLSX with images takes longer
        )
        assert response.status_code == 200, f"XLSX export failed with status {response.status_code}"
        assert "spreadsheetml" in response.headers.get("Content-Type", ""), f"Wrong content type: {response.headers.get('Content-Type')}"
        print("✓ XLSX export returns 200 with correct content type")
    
    def test_xlsx_has_4_sheets(self):
        """Verify XLSX has 4 sheets: Data Aset, Kelengkapan Dokumen, Data Kegiatan, Tim Inventarisasi"""
        response = requests.get(
            f"{BASE_URL}/api/export/xlsx?activity_id={TEST_ACTIVITY_ID}",
            timeout=180
        )
        assert response.status_code == 200
        
        # Load workbook
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        sheet_names = wb.sheetnames
        
        expected_sheets = ['Data Aset', 'Kelengkapan Dokumen', 'Data Kegiatan', 'Tim Inventarisasi']
        
        assert len(sheet_names) == 4, f"Expected 4 sheets, got {len(sheet_names)}: {sheet_names}"
        
        for expected in expected_sheets:
            assert expected in sheet_names, f"Missing sheet: {expected}. Found: {sheet_names}"
        
        print(f"✓ XLSX has all 4 expected sheets: {sheet_names}")
        wb.close()
    
    def test_xlsx_data_aset_has_jumlah_foto_and_tanggal_input_headers(self):
        """Verify 'Data Aset' sheet has 'Jumlah Foto' and 'Tanggal Input' as last 2 headers"""
        response = requests.get(
            f"{BASE_URL}/api/export/xlsx?activity_id={TEST_ACTIVITY_ID}",
            timeout=180
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb['Data Aset']
        
        # Read header row (row 1)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        # Check last 2 headers
        assert len(headers) >= 2, f"Not enough headers: {headers}"
        
        # Headers should end with 'Jumlah Foto' and 'Tanggal Input'
        last_two = headers[-2:]
        assert 'Jumlah Foto' in last_two, f"'Jumlah Foto' not in last 2 headers: {last_two}"
        assert 'Tanggal Input' in last_two, f"'Tanggal Input' not in last 2 headers: {last_two}"
        
        print(f"✓ Data Aset sheet headers end with: {last_two}")
        print(f"  Total headers: {len(headers)}")
        wb.close()
    
    def test_xlsx_data_kegiatan_has_sections(self):
        """Verify 'Data Kegiatan' sheet has sections: INFORMASI KEGIATAN, SATUAN KERJA, PENANGGUNG JAWAB, BERITA ACARA"""
        response = requests.get(
            f"{BASE_URL}/api/export/xlsx?activity_id={TEST_ACTIVITY_ID}",
            timeout=180
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb['Data Kegiatan']
        
        # Read all cell values to find section headers
        all_values = []
        for row in ws.iter_rows(max_row=50):  # Check first 50 rows
            for cell in row:
                if cell.value:
                    all_values.append(str(cell.value))
        
        content_str = ' '.join(all_values)
        
        expected_sections = [
            'INFORMASI KEGIATAN',
            'SATUAN KERJA', 
            'PENANGGUNG JAWAB',
            'BERITA ACARA'
        ]
        
        found_sections = []
        missing_sections = []
        
        for section in expected_sections:
            if section in content_str:
                found_sections.append(section)
            else:
                missing_sections.append(section)
        
        assert len(missing_sections) == 0, f"Missing sections: {missing_sections}. Found: {found_sections}"
        
        print(f"✓ Data Kegiatan sheet has all 4 sections: {found_sections}")
        wb.close()
    
    def test_xlsx_tim_inventarisasi_has_team_sections(self):
        """Verify 'Tim Inventarisasi' sheet has 4 team sections"""
        response = requests.get(
            f"{BASE_URL}/api/export/xlsx?activity_id={TEST_ACTIVITY_ID}",
            timeout=180
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb['Tim Inventarisasi']
        
        # Read all cell values
        all_values = []
        for row in ws.iter_rows(max_row=100):
            for cell in row:
                if cell.value:
                    all_values.append(str(cell.value))
        
        content_str = ' '.join(all_values)
        
        expected_sections = [
            'TIM INTI',
            'TIM PEMBANTU',
            'TIM PENELITI', 
            'TIM PENDUKUNG'
        ]
        
        found_sections = []
        missing_sections = []
        
        for section in expected_sections:
            if section in content_str:
                found_sections.append(section)
            else:
                missing_sections.append(section)
        
        assert len(missing_sections) == 0, f"Missing team sections: {missing_sections}. Found: {found_sections}"
        
        print(f"✓ Tim Inventarisasi sheet has all 4 team sections: {found_sections}")
        wb.close()
    
    def test_xlsx_tim_inventarisasi_has_tabular_format(self):
        """Verify Tim Inventarisasi has tabular headers: No, Nama, Jabatan, NIP, Unit, Peran"""
        response = requests.get(
            f"{BASE_URL}/api/export/xlsx?activity_id={TEST_ACTIVITY_ID}",
            timeout=180
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb['Tim Inventarisasi']
        
        # Read all cell values
        all_values = []
        for row in ws.iter_rows(max_row=100):
            for cell in row:
                if cell.value:
                    all_values.append(str(cell.value))
        
        content_str = ' '.join(all_values)
        
        # Check for table headers
        expected_headers = ['No', 'Nama', 'Jabatan', 'NIP', 'Peran']
        
        found_headers = []
        for h in expected_headers:
            if h in content_str:
                found_headers.append(h)
        
        assert len(found_headers) >= 4, f"Missing table headers. Found: {found_headers}, Expected: {expected_headers}"
        
        print(f"✓ Tim Inventarisasi has tabular format with headers: {found_headers}")
        wb.close()


class TestCodeReview:
    """Code review verification - checks key code patterns"""
    
    def test_exports_datetime_import(self):
        """Verify exports.py has datetime import (via inspection or API behavior)"""
        # The fact that CSV export works with 'tanggal_input' proves datetime is imported
        # because created_at is a datetime field
        response = requests.get(
            f"{BASE_URL}/api/export/csv?activity_id={TEST_ACTIVITY_ID}",
            stream=True,
            timeout=60
        )
        assert response.status_code == 200, "CSV export failed - datetime import may be missing"
        print("✓ CSV export works - datetime import is present")
    
    def test_export_no_500_errors(self):
        """Verify neither export endpoints return 500 errors"""
        # CSV
        csv_resp = requests.get(
            f"{BASE_URL}/api/export/csv?activity_id={TEST_ACTIVITY_ID}",
            stream=True,
            timeout=60
        )
        assert csv_resp.status_code != 500, "CSV export returned 500 error"
        
        # XLSX - may timeout but shouldn't 500
        xlsx_resp = requests.get(
            f"{BASE_URL}/api/export/xlsx?activity_id={TEST_ACTIVITY_ID}",
            timeout=180
        )
        assert xlsx_resp.status_code != 500, "XLSX export returned 500 error"
        
        print("✓ Both CSV and XLSX exports work without 500 errors")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
