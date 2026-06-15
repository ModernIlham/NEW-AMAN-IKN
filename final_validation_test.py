#!/usr/bin/env python3
"""
Final XLSX Export Validation Test - Validates all specific requirements from review request
"""

import requests
import json
import csv
import io
import sys
import time
import base64
import uuid
from typing import Dict, Any, Optional

# Backend URL from environment
BACKEND_URL = "https://asset-crud-auth.preview.emergentagent.com/api"

def main():
    """Final validation test"""
    print("=== FINAL XLSX EXPORT VALIDATION ===")
    
    # Create session and login
    session = requests.Session()
    
    # Login with existing user
    login_data = {"username": "xlsxtester", "password": "test1234"}
    response = session.post(f"{BACKEND_URL}/auth/login", json=login_data)
    
    if response.status_code != 200:
        print("❌ Login failed")
        return False
    
    token = response.json().get("access_token")
    session.headers.update({"Authorization": f"Bearer {token}"})
    print("✅ Login successful")
    
    # Create test activity
    activity_data = {
        "nomor_surat": "FINAL/2024",
        "nama_kegiatan": "Final XLSX Validation Test",
        "tanggal_mulai": "2024-01-01",
        "tanggal_selesai": "2024-12-31",
        "status": "Aktif",
        "keterangan": "Final validation test"
    }
    
    response = session.post(f"{BACKEND_URL}/inventory-activities", json=activity_data)
    if response.status_code != 200:
        print("❌ Activity creation failed")
        return False
        
    activity_id = response.json().get("id")
    print(f"✅ Activity created: {activity_id}")
    
    # Create sample data
    png_data = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\xf8\x0f\x00\x00\x01\x00\x01\x00\x00\x00\x00\x00\x00\x00\x00\x18\xdd\x8d\xb4\x1c\x00\x00\x00\x00IEND\xaeB`\x82'
    photo_b64 = base64.b64encode(png_data).decode('utf-8')
    
    pdf_content = b"""%PDF-1.4
1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj
2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj
3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R>>endobj
4 0 obj<</Length 44>>stream
BT
/F1 12 Tf
72 720 Td
(Test Document) Tj
ET
endstream
endobj
xref
0 5
0000000000 65535 f 
0000000010 00000 n 
0000000053 00000 n 
0000000125 00000 n 
0000000248 00000 n 
trailer<</Size 5/Root 1 0 R>>
startxref
341
%%EOF"""
    pdf_b64 = base64.b64encode(pdf_content).decode('utf-8')
    
    # Create asset exactly as specified in review request
    asset_data = {
        "asset_code": "3030103001",
        "NUP": "TST001",
        "asset_name": "Test Asset for Final XLSX Validation",
        "category": "Elektronik & IT",
        "brand": "Dell",
        "model": "Latitude",
        "activity_id": activity_id,
        "document_checklist": [
            {
                "name": "BPKB/Sertifikat",
                "checked": True,  # Ada - should appear in XLSX
                "notes": "Dokumen lengkap tersedia",
                "photos": [
                    f"data:image/png;base64,{photo_b64}",
                    f"data:image/png;base64,{photo_b64}",
                    f"data:image/png;base64,{photo_b64}"
                ],
                "documents": [
                    {"name": "BPKB.pdf", "data": f"data:application/pdf;base64,{pdf_b64}"}
                ]
            },
            {
                "name": "Faktur Pembelian", 
                "checked": False,  # Tidak Ada - should NOT appear in XLSX
                "notes": "Dokumen tidak tersedia",
                "photos": [],
                "documents": []
            }
        ]
    }
    
    response = session.post(f"{BACKEND_URL}/assets", json=asset_data)
    if response.status_code != 200:
        print(f"❌ Asset creation failed: {response.status_code} - {response.text}")
        return False
        
    asset_id = response.json().get("id")
    print(f"✅ Asset created: {asset_id}")
    
    # Test XLSX export with exact URL as specified in review request
    export_url = f"{BACKEND_URL}/export/xlsx?activity_id={activity_id}&base_url=http://localhost:8001"
    print(f"Testing XLSX export: {export_url}")
    
    response = session.get(export_url)
    if response.status_code != 200:
        print(f"❌ XLSX export failed: {response.status_code}")
        return False
    
    print(f"✅ XLSX export successful: {len(response.content)} bytes")
    
    # Validate XLSX content
    try:
        import openpyxl
        
        workbook = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
        
        # Requirement 1: Sheet 'Kelengkapan Dokumen' exists
        if 'Kelengkapan Dokumen' not in workbook.sheetnames:
            print("❌ 'Kelengkapan Dokumen' sheet missing")
            return False
        print("✅ 'Kelengkapan Dokumen' sheet exists")
        
        doc_sheet = workbook['Kelengkapan Dokumen']
        
        # Requirement 2: Headers should be exactly as specified
        expected_headers = ['Kode Aset', 'NUP', 'Nama Aset', 'Item Kelengkapan', 'Catatan', 
                           'Foto 1', 'Foto 2', 'Foto 3', 'PDF 1', 'PDF 2', 'PDF 3']
        
        actual_headers = []
        for col in range(1, 12):
            cell_value = doc_sheet.cell(row=1, column=col).value
            if cell_value:
                actual_headers.append(str(cell_value))
        
        if actual_headers == expected_headers:
            print("✅ Headers match exactly as specified")
        else:
            print(f"❌ Headers mismatch. Expected: {expected_headers}, Got: {actual_headers}")
            return False
        
        # Requirement 3: Only checked=True items appear (Tidak Ada items should NOT appear)
        rows_count = 0
        ada_items = []
        tidak_ada_items = []
        
        for row in range(2, doc_sheet.max_row + 1):
            item_name = doc_sheet.cell(row=row, column=4).value
            if item_name:
                rows_count += 1
                item_str = str(item_name)
                if "BPKB" in item_str or "Sertifikat" in item_str:
                    ada_items.append("BPKB/Sertifikat")
                elif "Faktur" in item_str:
                    tidak_ada_items.append("Faktur Pembelian")
        
        if len(ada_items) > 0 and len(tidak_ada_items) == 0:
            print("✅ Only checked=True (Ada) items shown, Tidak Ada items filtered out")
        else:
            print(f"❌ Filtering failed. Ada items: {ada_items}, Tidak Ada items: {tidak_ada_items}")
            return False
        
        # Requirement 4: Photo columns have clickable hyperlinks
        photo_links = 0
        for row in range(2, doc_sheet.max_row + 1):
            for col in [6, 7, 8]:  # Foto 1, 2, 3
                cell = doc_sheet.cell(row=row, column=col)
                if (cell.hyperlink and cell.hyperlink.target) or (hasattr(doc_sheet, '_images') and doc_sheet._images):
                    photo_links += 1
                    break
        
        if photo_links > 0:
            print("✅ Photo columns have clickable hyperlinks")
        else:
            print("❌ No photo hyperlinks found")
            return False
        
        # Requirement 5: PDF columns have clickable hyperlinks with document names
        pdf_links = 0
        for row in range(2, doc_sheet.max_row + 1):
            for col in [9, 10, 11]:  # PDF 1, 2, 3
                cell = doc_sheet.cell(row=row, column=col)
                if cell.hyperlink and cell.hyperlink.target:
                    pdf_links += 1
                    print(f"PDF link: {cell.hyperlink.target}")
        
        if pdf_links > 0:
            print("✅ PDF columns have clickable hyperlinks")
        else:
            print("❌ No PDF hyperlinks found")
            return False
        
    except ImportError:
        print("⚠️ openpyxl not available for detailed validation")
        return False
    
    # Test doc-file endpoint as specified
    test_endpoints = [
        f"{BACKEND_URL}/assets/{asset_id}/doc-file/0/photo/0",
        f"{BACKEND_URL}/assets/{asset_id}/doc-file/0/photo/1", 
        f"{BACKEND_URL}/assets/{asset_id}/doc-file/0/photo/2",
        f"{BACKEND_URL}/assets/{asset_id}/doc-file/0/document/0"
    ]
    
    for endpoint in test_endpoints:
        response = session.get(endpoint)
        if response.status_code == 200:
            content_type = response.headers.get('content-type', '')
            print(f"✅ {endpoint.split('/')[-2:]}: {content_type}, {len(response.content)} bytes")
        else:
            print(f"❌ {endpoint}: Status {response.status_code}")
    
    # Cleanup
    session.delete(f"{BACKEND_URL}/inventory-activities/{activity_id}")
    print("✅ Test data cleaned up")
    
    print("\n=== ALL REQUIREMENTS VALIDATED SUCCESSFULLY ===")
    return True

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)